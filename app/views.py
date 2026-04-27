import hashlib
import io
import json
import logging
import re
import time
import uuid
from calendar import monthrange
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import (
    authenticate,
    get_user_model,
    login,
    logout,
    update_session_auth_hash,
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth.tokens import default_token_generator
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import Avg, Count, Exists, Max, OuterRef, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.templatetags.static import static
from django.urls import reverse
from django.utils import timezone
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.views.decorators.http import require_GET, require_POST
from django.views.generic import TemplateView
from xhtml2pdf import pisa

from .models import (
    AboutContent,
    AboutValueItem,
    AdditionalOnly,
    AddOn,
    AdminNotification,
    AuditLog,
    Booking,
    BookingImage,
    CanvasAsset,
    CanvasCategory,
    CanvasLabel,
    ChatMessage,
    ChatSession,
    ConcernTicket,
    Design,
    GalleryCategory,
    GalleryImage,
    GCashConfig,
    HomeContent,
    HomeFeatureItem,
    Notification,
    Package,
    Payment,
    Review,
    ReviewImage,
    Service,
    ServiceChargeConfig,
    ServiceContent,
    User,
    UserDesign,
)
from .services import (
    create_paymongo_checkout_session,
    get_current_ban_status,
    get_chatbot_response,
    retrieve_paymongo_checkout_session,
    retrieve_paymongo_payment,
    verify_paymongo_webhook_signature,
)

logger = logging.getLogger(__name__)


def _increase_rate_limit_counter(key, timeout_seconds):
    if cache.add(key, 1, timeout=timeout_seconds):
        return 1
    try:
        return cache.incr(key)
    except ValueError:
        cache.set(key, 1, timeout=timeout_seconds)
        return 1


def _get_client_ip(request):
    forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "unknown")


def _format_wait_time(seconds):
    seconds = max(1, int(seconds))
    if seconds < 60:
        return f"{seconds} second{'s' if seconds != 1 else ''}"
    minutes, rem_seconds = divmod(seconds, 60)
    if minutes < 60:
        if rem_seconds:
            return f"{minutes} minute{'s' if minutes != 1 else ''} and {rem_seconds} second{'s' if rem_seconds != 1 else ''}"
        return f"{minutes} minute{'s' if minutes != 1 else ''}"
    hours, rem_minutes = divmod(minutes, 60)
    if rem_minutes:
        return f"{hours} hour{'s' if hours != 1 else ''} and {rem_minutes} minute{'s' if rem_minutes != 1 else ''}"
    return f"{hours} hour{'s' if hours != 1 else ''}"


def _is_reset_request_rate_limited(request, email):
    client_ip = _get_client_ip(request)
    normalized_email = (email or "").strip().lower()
    email_hash = hashlib.sha256(normalized_email.encode("utf-8")).hexdigest()
    now = int(time.time())

    cooldown_seconds = getattr(settings, "FORGOT_PASSWORD_COOLDOWN_SECONDS", 60)
    window_seconds = getattr(
        settings, "FORGOT_PASSWORD_RATE_LIMIT_WINDOW_SECONDS", 3600
    )
    ip_limit = getattr(settings, "FORGOT_PASSWORD_RATE_LIMIT_PER_IP", 5)
    email_limit = getattr(settings, "FORGOT_PASSWORD_RATE_LIMIT_PER_EMAIL", 3)

    cooldown_key = f"pwd-reset:cooldown:ip:{client_ip}"
    cooldown_until_key = f"{cooldown_key}:until"
    ip_count_key = f"pwd-reset:count:ip:{client_ip}"
    ip_window_until_key = f"{ip_count_key}:until"
    email_count_key = f"pwd-reset:count:email:{email_hash}"
    email_window_until_key = f"{email_count_key}:until"

    if cache.get(cooldown_key):
        cooldown_until = cache.get(cooldown_until_key) or (now + cooldown_seconds)
        wait_seconds = max(1, int(cooldown_until) - now)
        return (
            True,
            f"Please wait {_format_wait_time(wait_seconds)} before requesting another reset link.",
        )

    current_ip_count = cache.get(ip_count_key, 0)
    current_email_count = cache.get(email_count_key, 0)
    if current_ip_count >= ip_limit or current_email_count >= email_limit:
        ip_wait_seconds = 0
        email_wait_seconds = 0

        if current_ip_count >= ip_limit:
            ip_window_until = cache.get(ip_window_until_key) or (now + window_seconds)
            ip_wait_seconds = max(1, int(ip_window_until) - now)

        if current_email_count >= email_limit:
            email_window_until = cache.get(email_window_until_key) or (
                now + window_seconds
            )
            email_wait_seconds = max(1, int(email_window_until) - now)

        wait_seconds = max(ip_wait_seconds, email_wait_seconds, 1)
        return (
            True,
            f"Too many reset attempts. Please try again in {_format_wait_time(wait_seconds)}.",
        )

    cache.set(cooldown_key, 1, timeout=cooldown_seconds)
    cache.set(cooldown_until_key, now + cooldown_seconds, timeout=cooldown_seconds)
    cache.add(ip_window_until_key, now + window_seconds, timeout=window_seconds)
    cache.add(email_window_until_key, now + window_seconds, timeout=window_seconds)
    _increase_rate_limit_counter(ip_count_key, window_seconds)
    _increase_rate_limit_counter(email_count_key, window_seconds)
    return False, ""


def _send_account_verification_email(request, user):
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    verify_url = request.build_absolute_uri(
        reverse("verify_email", kwargs={"uidb64": uid, "token": token})
    )
    subject = "Verify your Balloorina account"
    body = (
        f"Hi {user.first_name or user.username},\n\n"
        f"Thanks for registering at Balloorina.\n"
        f"Please verify your Gmail by clicking the link below:\n{verify_url}\n\n"
        f"If you did not create this account, you can ignore this email."
    )
    from_email = getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@balloorina.local")
    send_mail(
        subject,
        body,
        from_email,
        [user.email],
        fail_silently=False,
    )


def log_action(user, action):
    """Helper function to create an audit log entry."""
    AuditLog.objects.create(user=user, action=action)


def get_service_charge_config():
    config, _ = ServiceChargeConfig.objects.get_or_create(
        id=1,
        defaults={
            "amount": Decimal("0.00"),
            "notes": "Includes styling fee, toll fees, fuel, crew meals, and ingress/egress logistics.",
        },
    )
    return config


def normalize_package_part(value):
    cleaned = re.sub(
        r"\s*\((add-on|additional|solo)\)\s*$", "", value or "", flags=re.IGNORECASE
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned.lower()


def format_booking_selection(package_type):
    parts = [p.strip() for p in (package_type or "").split("+") if p.strip()]
    if not parts:
        return "Custom Booking"

    packages = []
    addons = []
    solo_addons = []
    additionals = []

    for part in parts:
        if re.search(r"\(solo\)\s*$", part, re.IGNORECASE):
            solo_addons.append(
                re.sub(r"\s*\(solo\)\s*$", "", part, flags=re.IGNORECASE).strip()
            )
        elif re.search(r"\(add-on\)\s*$", part, re.IGNORECASE):
            addons.append(
                re.sub(r"\s*\(add-on\)\s*$", "", part, flags=re.IGNORECASE).strip()
            )
        elif re.search(r"\(additional\)\s*$", part, re.IGNORECASE):
            additionals.append(
                re.sub(r"\s*\(additional\)\s*$", "", part, flags=re.IGNORECASE).strip()
            )
        else:
            packages.append(part)

    segments = []
    if packages:
        segments.append(f"Package: {', '.join(packages)}")
    if addons:
        segments.append(f"Add-on: {', '.join(addons)}")
    if solo_addons:
        segments.append(f"Solo Add-on: {', '.join(solo_addons)}")
    if additionals:
        segments.append(f"Additional: {', '.join(additionals)}")

    return " | ".join(segments) if segments else "Custom Booking"


def get_booking_price_breakdown(booking):
    parts = [p.strip() for p in (booking.package_type or "").split("+") if p.strip()]
    breakdown = []
    subtotal = Decimal("0.00")

    package_map = {
        normalize_package_part(pkg.name): pkg for pkg in Package.objects.all()
    }
    addon_map = {
        normalize_package_part(addon.name): addon for addon in AddOn.objects.all()
    }
    additional_map = {
        normalize_package_part(additional.name): additional
        for additional in AdditionalOnly.objects.all()
    }

    for raw_part in parts:
        is_solo = raw_part.endswith("(Solo)")
        part_name = raw_part.replace("(Solo)", "").strip() if is_solo else raw_part
        normalized_part_name = normalize_package_part(part_name)
        amount = None
        label = raw_part

        if is_solo:
            addon = addon_map.get(normalized_part_name)
            if addon and addon.solo_price is not None:
                amount = addon.solo_price
                label = f"{addon.name.strip()} (Solo)"
        else:
            package_obj = package_map.get(normalized_part_name)
            if package_obj:
                amount = package_obj.price
                label = package_obj.name.strip()
            else:
                addon = addon_map.get(normalized_part_name)
                if addon:
                    amount = addon.price
                    label = f"{addon.name.strip()} (Add-on)"
                else:
                    additional = additional_map.get(normalized_part_name)
                    if additional:
                        amount = additional.price
                        label = f"{additional.name.strip()} (Additional)"

        if amount is not None:
            subtotal += amount
            breakdown.append({"label": label, "amount": amount})

    service_charge = Decimal("0.00")
    if parts:
        service_charge = get_service_charge_config().amount or Decimal("0.00")

    computed_total = subtotal + service_charge
    return {
        "items": breakdown,
        "subtotal": subtotal,
        "service_charge": service_charge,
        "computed_total": computed_total,
        "stored_total": booking.total_price or Decimal("0.00"),
    }


def build_booking_snapshot(booking):
    return {
        "event_type": booking.event_type or "",
        "event_date": booking.event_date.isoformat() if booking.event_date else "",
        "event_time": booking.event_time.strftime("%H:%M")
        if booking.event_time
        else "",
        "event_location": booking.event_location or "",
        "package_type": booking.package_type or "",
        "special_requests": booking.special_requests or "",
        "total_price": str(booking.total_price or Decimal("0.00")),
    }


def apply_booking_snapshot(booking, snapshot):
    if not snapshot:
        return

    booking.event_type = snapshot.get("event_type", booking.event_type)
    event_date_raw = snapshot.get("event_date")
    event_time_raw = snapshot.get("event_time")
    total_price_raw = snapshot.get("total_price")

    if event_date_raw:
        try:
            booking.event_date = datetime.strptime(event_date_raw, "%Y-%m-%d").date()
        except ValueError:
            pass

    if event_time_raw:
        try:
            booking.event_time = datetime.strptime(event_time_raw, "%H:%M").time()
        except ValueError:
            booking.event_time = None
    else:
        booking.event_time = None

    booking.event_location = snapshot.get("event_location", booking.event_location)
    booking.package_type = snapshot.get("package_type", booking.package_type)
    booking.special_requests = snapshot.get(
        "special_requests", booking.special_requests
    )

    if total_price_raw:
        try:
            booking.total_price = Decimal(str(total_price_raw))
        except (InvalidOperation, TypeError):
            pass


def get_top_reviews():
    """Helper function to get the top 3 reviews, prioritizing 5-stars and unique per user."""
    # Fetch reviews ordered by rating descending, then newest first
    all_reviews = Review.objects.select_related("user", "booking").order_by(
        "-rating", "-created_at"
    )

    top_reviews = []
    seen_users = set()

    for review in all_reviews:
        if review.user.id not in seen_users:
            top_reviews.append(review)
            seen_users.add(review.user.id)

        if len(top_reviews) >= 6:
            break

    for review in top_reviews:
        review.booking_selection_display = format_booking_selection(
            review.booking.package_type if review.booking else ""
        )

    return top_reviews


class HomePageView(TemplateView):
    template_name = "client/home.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["home_content"] = HomeContent.objects.first()
        context["home_features"] = HomeFeatureItem.objects.filter(is_active=True)
        context["top_reviews"] = get_top_reviews()
        context["latest_creations"] = GalleryImage.objects.filter(
            is_active=True
        ).order_by("-id")[:6]
        return context


class AboutPageView(TemplateView):
    template_name = "client/about.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["about_content"] = AboutContent.objects.first()
        context["about_values"] = AboutValueItem.objects.filter(is_active=True)
        context["top_reviews"] = get_top_reviews()
        return context


class ServicesPageView(TemplateView):
    template_name = "client/services.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["service_content"] = ServiceContent.objects.first()
        context["services"] = Service.objects.filter(is_active=True).order_by("display_order")
        return context


class GuidelinesPageView(TemplateView):
    template_name = "client/guidelines.html"


class PackagePageView(TemplateView):
    template_name = "client/package.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        service_charge_config = get_service_charge_config()
        context.update(
            {
                "packages": Package.objects.all().order_by(
                    "-is_featured", "-created_at"
                ),
                "addons": AddOn.objects.filter(is_active=True).order_by("-created_at"),
                "additionals": AdditionalOnly.objects.filter(is_active=True).order_by(
                    "-created_at"
                ),
                "service_charge_amount": service_charge_config.amount,
                "service_charge_notes": service_charge_config.notes,
            }
        )
        return context


class GalleryPageView(TemplateView):
    template_name = "client/gallery.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["categories"] = GalleryCategory.objects.all()
        context["gallery_images"] = GalleryImage.objects.filter(
            is_active=True
        ).select_related("category")
        return context


def reviews_page(request):
    reviews = Review.objects.select_related("user", "booking").order_by("-created_at")

    import json

    # Check if the current user has liked each review
    if request.user.is_authenticated:
        for review in reviews:
            review.is_liked_by_user = review.likes.filter(id=request.user.id).exists()
            review.can_be_liked = request.user != review.user

            # Serialize images for editing if the user owns the review
            if review.user == request.user:
                images_data = [
                    {"id": img.id, "url": img.image.url} for img in review.images.all()
                ]
                review.images_json = json.dumps(images_data)
    else:
        for review in reviews:
            review.is_liked_by_user = False
            review.can_be_liked = False

    return render(request, "client/reviews.html", {"reviews": reviews})


@login_required
@require_POST
def like_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)

    # Prevent user from liking their own review
    if review.user == request.user:
        return JsonResponse({"error": "You cannot like your own review."}, status=400)

    # Toggle like
    if review.likes.filter(id=request.user.id).exists():
        review.likes.remove(request.user)
        liked = False
    else:
        review.likes.add(request.user)
        liked = True

    return JsonResponse({"liked": liked, "total_likes": review.total_likes()})


@login_required
@require_POST
def edit_review(request, review_id):
    review = get_object_or_404(Review, id=review_id, user=request.user)

    rating = request.POST.get("rating")
    comment = request.POST.get("comment")
    images_to_delete = request.POST.getlist("delete_images[]")
    new_images = request.FILES.getlist("images")

    # Validate rating range
    try:
        rating_val = int(rating) if rating else 0
    except (ValueError, TypeError):
        rating_val = 0
    if rating_val < 1 or rating_val > 5:
        return JsonResponse(
            {"status": "error", "message": "Rating must be between 1 and 5."},
            status=400,
        )

    if rating and comment:
        # Calculate resulting image count
        current_images_count = review.images.count()
        resulting_count = current_images_count - len(images_to_delete) + len(new_images)

        if resulting_count > 4:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "You can only have a maximum of 4 pictures per review.",
                },
                status=400,
            )

        # 1. Delete requested images
        if images_to_delete:
            for img_id in images_to_delete:
                try:
                    img = ReviewImage.objects.get(id=img_id, review=review)
                    img.image.delete()  # Deletes file from storage
                    img.delete()  # Deletes record from DB
                except ReviewImage.DoesNotExist:
                    pass

        # 2. Add new images
        if new_images:
            for image in new_images:
                ReviewImage.objects.create(review=review, image=image)

        # 3. Update Text and Rating
        review.rating = rating
        review.comment = comment
        review.save()

        log_action(request.user, f"Updated review #{review.id}.")
        return JsonResponse(
            {
                "status": "success",
                "message": "Review updated successfully!",
                "rating": review.rating,
                "comment": review.comment,
            }
        )

    return JsonResponse(
        {"status": "error", "message": "Rating and comment are required."}, status=400
    )


@login_required
@require_POST
def delete_review(request, review_id):
    review = get_object_or_404(Review, id=review_id, user=request.user)
    review_id_val = review.id
    review.delete()

    log_action(request.user, f"Deleted review #{review_id_val}.")
    return JsonResponse(
        {"status": "success", "message": "Review deleted successfully!"}
    )


User = get_user_model()


def register(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        phone = request.POST.get("phone", "").strip()
        role = "customer"

        errors = []

        # Required fields
        if not first_name:
            errors.append("First name is required.")
        if not last_name:
            errors.append("Last name is required.")
        if not username:
            errors.append("Username is required.")
        if not email:
            errors.append("Email is required.")
        elif not re.fullmatch(r"[A-Za-z0-9._%+-]+@gmail\.com", email):
            errors.append(
                "Please enter a valid Gmail address (example: yourname@gmail.com)."
            )
        if not password:
            errors.append("Password is required.")
        if not confirm_password:
            errors.append("Confirm password is required.")
        if not phone:
            errors.append("Phone number is required.")

        # Email unique
        if User.objects.filter(email=email).exists():
            errors.append("Email already exists.")

        # Username unique
        if User.objects.filter(username=username).exists():
            errors.append("Username already exists.")

        # Password match
        if password != confirm_password:
            errors.append("Passwords do not match.")

        # Password rules
        if len(password) < 8:
            errors.append("Password must be at least 8 characters.")

        # Phone number validation
        cleaned_phone = re.sub(r"[\s\-\(\)\+]", "", phone)
        if phone and not cleaned_phone.isdigit():
            errors.append("Phone number must contain only digits.")
        elif phone and (len(cleaned_phone) < 10 or len(cleaned_phone) > 15):
            errors.append("Phone number must be between 10 and 15 digits.")

        if errors:
            return render(request, "auth/register.html", {"errors": errors})

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    phone_number=phone,
                    role=role,
                    email_verified=False,
                )
                _send_account_verification_email(request, user)
        except Exception:
            logger.exception("Failed to send account verification email.")
            errors.append(
                "We could not send a verification email right now. Please try again in a moment."
            )
            return render(request, "auth/register.html", {"errors": errors})

        log_action(None, f"New user '{username}' registered. Verification email sent.")
        messages.success(
            request,
            "Registration successful! Please check your Gmail and verify your account before logging in.",
        )
        return redirect("login")

    return render(request, "auth/register.html")


def verify_email(request, uidb64, token):
    user = None
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if not user:
        messages.error(request, "Invalid verification link.")
        return redirect("login")

    if user.email_verified:
        messages.success(request, "Your email is already verified. You can log in.")
        return redirect("login")

    if not default_token_generator.check_token(user, token):
        messages.error(
            request,
            "This verification link is invalid or expired. Please register again.",
        )
        return redirect("register")

    user.email_verified = True
    user.save(update_fields=["email_verified"])
    log_action(user, "Email verified.")
    messages.success(request, "Email verified successfully. You can now log in.")
    return redirect("login")


User = get_user_model()


def user_login(request):
    if request.method == "POST":
        email = (request.POST.get("email") or "").strip().lower()
        password = request.POST.get("password")

        # Try to get user by email
        try:
            user_obj = User.objects.get(email=email)
            username = user_obj.username  # Django authenticate needs username
        except User.DoesNotExist:
            return render(
                request, "auth/login.html", {"error": "Invalid email or password."}
            )

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if not getattr(user, "email_verified", True):
                return render(
                    request,
                    "auth/login.html",
                    {
                        "error": "Please verify your Gmail first. Check your inbox for the verification link."
                    },
                )

            login(request, user)

            log_action(user, "User logged in.")
            if request.POST.get("remember_me"):
                request.session.set_expiry(1209600)  # 2 weeks
            else:
                request.session.set_expiry(0)  # browser close

            if user.role == "customer":
                messages.success(request, "Login successful! Welcome.")
                return redirect("home")
            elif user.role in ["admin", "staff"]:
                messages.success(request, "Login successful! Welcome to the dashboard.")
                return redirect("dashboard")
        else:
            return render(
                request, "auth/login.html", {"error": "Invalid email or password."}
            )

    return render(request, "auth/login.html")


def forgot_password_request(request):
    if request.method == "POST":
        email = (request.POST.get("email") or "").strip()
        generic_message = "If the email exists, a password reset link has been sent."

        if not email:
            messages.error(request, "Please enter your account email.")
            return render(request, "auth/forgot_password.html")

        is_limited, rate_limit_message = _is_reset_request_rate_limited(request, email)
        if is_limited:
            messages.error(request, rate_limit_message)
            return render(request, "auth/forgot_password.html")

        try:
            user = User.objects.get(email__iexact=email)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            reset_url = request.build_absolute_uri(
                reverse(
                    "password_reset_confirm", kwargs={"uidb64": uid, "token": token}
                )
            )
            subject = "Balloorina Password Reset"
            body = (
                f"Hi {user.first_name or user.username},\n\n"
                f"We received a request to reset your password.\n"
                f"Use the link below:\n{reset_url}\n\n"
                f"If you did not request this, you can ignore this message."
            )
            from_email = getattr(
                settings, "DEFAULT_FROM_EMAIL", "no-reply@balloorina.local"
            )
            send_mail(
                subject,
                body,
                from_email,
                [user.email],
                fail_silently=getattr(settings, "EMAIL_FAIL_SILENTLY", True),
            )
        except User.DoesNotExist:
            pass
        except Exception:
            logger.exception("Failed to send forgot-password email.")

        messages.success(request, generic_message)
        return redirect("forgot_password")

    return render(request, "auth/forgot_password.html")


def password_reset_confirm(request, uidb64, token):
    user = None
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    is_valid_link = bool(user and default_token_generator.check_token(user, token))

    if request.method == "POST":
        if not is_valid_link:
            messages.error(request, "This reset link is invalid or expired.")
            return redirect("forgot_password")

        password = request.POST.get("password") or ""
        confirm_password = request.POST.get("confirm_password") or ""

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(
                request, "auth/reset_password.html", {"is_valid_link": is_valid_link}
            )

        try:
            validate_password(password, user=user)
        except ValidationError as exc:
            messages.error(request, " ".join(exc.messages))
            return render(
                request, "auth/reset_password.html", {"is_valid_link": is_valid_link}
            )

        user.set_password(password)
        user.save()
        log_action(user, "Password reset via forgot password.")
        messages.success(request, "Password updated successfully. You can now log in.")
        return redirect("login")

    return render(request, "auth/reset_password.html", {"is_valid_link": is_valid_link})


@login_required
def report_concern(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    form_error_message = ""
    form_error_toasts = []
    form_data = {
        "category": "",
        "other_category": "",
        "subject": "",
        "message": "",
    }

    if request.method == "POST":
        category = (request.POST.get("category") or "").strip()
        subject = (request.POST.get("subject") or "").strip()
        message_text = (request.POST.get("message") or "").strip()
        other_category = (request.POST.get("other_category") or "").strip()

        form_data = {
            "category": category,
            "other_category": other_category,
            "subject": subject,
            "message": message_text,
        }

        valid_categories = {choice[0] for choice in ConcernTicket.CATEGORY_CHOICES}

        if category not in valid_categories:
            form_error_message = "Please select a valid category."
        elif category == "other" and not other_category:
            form_error_message = "Please specify your concern type."
        elif not subject:
            form_error_message = "Please enter a subject."
        elif not message_text:
            form_error_message = "Please enter your message."
        else:
            if category == "other" and other_category:
                subject = f"{subject} (Other: {other_category})"
            ticket = ConcernTicket.objects.create(
                user=request.user,
                category=category,
                subject=subject,
                message=message_text,
            )
            log_action(request.user, f"Submitted concern ticket #{ticket.id}.")
            messages.success(
                request, "Concern submitted. Our team will review it soon."
            )
            return redirect("report_concern")

    if form_error_message:
        form_error_toasts = [form_error_message]

    my_tickets = ConcernTicket.objects.filter(user=request.user).order_by(
        "-created_at"
    )[:20]
    return render(
        request,
        "client/report_concern.html",
        {
            "my_tickets": my_tickets,
            "form_data": form_data,
            "form_error_toasts": form_error_toasts,
        },
    )


def user_logout(request):
    if request.user.is_authenticated:
        log_action(request.user, "User logged out.")
    logout(request)
    return redirect("home")


@login_required
def change_password(request):
    if request.method == "POST":
        # Check if it's an AJAX request
        is_ajax = request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest"

        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            log_action(request.user, "Changed their password.")
            update_session_auth_hash(
                request, user
            )  # Important to keep the user logged in
            if is_ajax:
                return JsonResponse(
                    {"success": True, "message": "Password updated successfully!"}
                )
            messages.success(request, "Your password was successfully updated!")
        else:
            if is_ajax:
                return JsonResponse({"success": False, "errors": form.errors})
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
    return redirect("customer_profile")


@login_required
def dashboard(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    return render(request, "admin/dashboard.html", build_dashboard_context(request))


def check_booking_expirations():
    """Find pending bookings past their event date, mark as expired and notify user."""
    expired_bookings = Booking.objects.filter(
        status="pending", event_date__lt=timezone.now().date()
    )
    for b in expired_bookings:
        b.status = "expired"
        b.save()
        Notification.objects.create(
            user=b.user,
            booking=b,
            message=f"Your booking #{b.id} for {b.event_date} has expired because it was not confirmed in time.",
        )


@login_required
def my_profile(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    user_bookings = Booking.objects.filter(user=request.user)
    total_bookings = user_bookings.count()
    pending_count = user_bookings.filter(status="pending").count()
    confirmed_count = user_bookings.filter(status="confirmed").count()
    completed_count = user_bookings.filter(status="completed").count()

    return render(
        request,
        "client/my_profile.html",
        {
            "total_bookings": total_bookings,
            "pending_count": pending_count,
            "confirmed_count": confirmed_count,
            "completed_count": completed_count,
        },
    )


@login_required
def my_reviews(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    reviews_list = (
        Review.objects.filter(user=request.user)
        .select_related("booking")
        .order_by("-created_at")
    )

    # Pagination for published reviews (5 per page)
    paginator_published = Paginator(reviews_list, 5)
    page_published_number = request.GET.get("page_published", 1)
    try:
        reviews = paginator_published.page(page_published_number)
    except PageNotAnInteger:
        reviews = paginator_published.page(1)
    except EmptyPage:
        reviews = paginator_published.page(paginator_published.num_pages)

    # Check if the current user has liked each review (though they are their own reviews, just in case template expects it)
    for review in reviews:
        review.is_liked_by_user = review.likes.filter(id=request.user.id).exists()
        review.can_be_liked = False  # cannot like own review

        images_data = [
            {"id": img.id, "url": img.image.url} for img in review.images.all()
        ]
        review.images_json = json.dumps(images_data)

    # Fetch completed bookings without reviews
    pending_reviews_list = Booking.objects.filter(
        user=request.user, status="completed", reviews__isnull=True
    ).order_by("-event_date")

    # Pagination for pending reviews (5 per page)
    paginator_pending = Paginator(pending_reviews_list, 5)
    page_pending_number = request.GET.get("page_pending", 1)
    try:
        pending_reviews = paginator_pending.page(page_pending_number)
    except PageNotAnInteger:
        pending_reviews = paginator_pending.page(1)
    except EmptyPage:
        pending_reviews = paginator_pending.page(paginator_pending.num_pages)

    # Attach formatted time range
    for b in pending_reviews:
        b.time_range_display = get_booking_time_range(b)

    return render(
        request,
        "client/my_reviews.html",
        {
            "reviews": reviews,
            "published_page_obj": reviews,
            "pending_reviews": pending_reviews,
            "pending_page_obj": pending_reviews,
        },
    )


@login_required
def customer_profile(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    _sync_customer_booking_payment_states(request.user)

    # Auto-expire pending bookings in the past
    check_booking_expirations()

    # Get filter parameters
    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "all")
    sort_date = request.GET.get("sort_date", "desc")

    user_bookings = Booking.objects.filter(user=request.user)

    # Calculate Stats before filtering so overall stats remain correct
    total_bookings = user_bookings.count()
    pending_count = user_bookings.filter(status="pending").count()
    confirmed_count = user_bookings.filter(status="confirmed").count()
    completed_count = user_bookings.filter(status="completed").count()

    # Apply Search Filter (by ID or Event Type)
    if search_query:
        if search_query.isdigit():
            user_bookings = user_bookings.filter(
                Q(id=search_query) | Q(event_type__icontains=search_query)
            )
        else:
            user_bookings = user_bookings.filter(event_type__icontains=search_query)

    # Apply Status Filter
    if status_filter != "all":
        if status_filter == "request_edit":
            user_bookings = user_bookings.filter(
                status="confirmed", edit_requested=True
            )
        elif status_filter == "request_cancel":
            user_bookings = user_bookings.filter(status="cancel_requested")
        else:
            user_bookings = user_bookings.filter(status=status_filter)

    # Apply Sorting
    if sort_date == "id_desc":
        user_bookings = user_bookings.order_by("-id")
    elif sort_date == "id_asc":
        user_bookings = user_bookings.order_by("id")
    elif sort_date == "oldest":
        user_bookings = user_bookings.order_by("event_date")
    else:
        user_bookings = user_bookings.order_by("-event_date")

    # Pagination
    paginator = Paginator(user_bookings, 10)  # Show 10 bookings per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Attach formatted time range for the table/modal and check if reviewed
    for b in page_obj:
        b.time_range_display = get_booking_time_range(b)
        if b.status == "completed":
            b.has_reviewed = b.reviews.filter(user=request.user).exists()
        else:
            b.has_reviewed = False
        b.price_breakdown = get_booking_price_breakdown(b)

    # Get active packages, addons, additionals for the edit modal dropdown
    active_packages = Package.objects.filter(is_active=True)
    active_addons = AddOn.objects.filter(is_active=True)
    active_additionals = AdditionalOnly.objects.all()
    service_charge_config = get_service_charge_config()

    return render(
        request,
        "client/customer_profile.html",
        {
            "page_obj": page_obj,
            "user_bookings": page_obj,  # To maintain some backward compatibility for template logic, though page_obj is better
            "total_bookings": total_bookings,
            "pending_count": pending_count,
            "confirmed_count": confirmed_count,
            "completed_count": completed_count,
            "packages": active_packages,
            "active_addons": active_addons,
            "active_additionals": active_additionals,
            "global_service_charge": service_charge_config.amount,
            "global_service_charge_note": service_charge_config.notes,
            "search_query": search_query,
            "status_filter": status_filter,
            "sort_date": sort_date,
        },
    )


def _sync_customer_booking_payment_states(user):
    zero = Decimal("0.00")
    bookings = Booking.objects.filter(
        user=user, status__in=["pending_payment", "confirmed"]
    ).annotate(
        verified_total=Coalesce(
            Sum("payments__amount", filter=Q(payments__payment_status="verified")),
            Decimal("0.00"),
        )
    )

    now = timezone.now()
    to_update = []

    for booking in bookings:
        verified_total = booking.verified_total or zero
        total_price = booking.total_price or zero

        if verified_total >= total_price and total_price > zero:
            target_payment_status = "paid"
        elif verified_total > zero:
            target_payment_status = "partial"
        else:
            target_payment_status = "pending"

        target_status = booking.status
        if booking.status == "pending_payment" and verified_total > zero:
            target_status = "confirmed"

        if (
            booking.payment_status != target_payment_status
            or booking.status != target_status
        ):
            booking.payment_status = target_payment_status
            booking.status = target_status
            booking.updated_at = now
            to_update.append(booking)

    if to_update:
        Booking.objects.bulk_update(
            to_update, ["payment_status", "status", "updated_at"]
        )


@login_required
@require_POST
def submit_review(request, id):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    booking = get_object_or_404(Booking, id=id, user=request.user)

    referer = request.META.get("HTTP_REFERER", "")
    fallback_redirect = "customer_profile"
    if "my-reviews" in referer or "my_reviews" in referer:
        fallback_redirect = "my_reviews"

    if booking.status != "completed":
        messages.error(request, "You can only review completed bookings.")
        return redirect(fallback_redirect)

    # Check if already reviewed
    if booking.reviews.filter(user=request.user).exists():
        messages.error(request, "You have already reviewed this booking.")
        return redirect(fallback_redirect)

    rating = request.POST.get("rating")
    comment = request.POST.get("comment")

    if rating and comment:
        # Validate rating range
        try:
            rating_val = int(rating)
        except (ValueError, TypeError):
            messages.error(request, "Invalid rating value.")
            return redirect(fallback_redirect)
        if rating_val < 1 or rating_val > 5:
            messages.error(request, "Rating must be between 1 and 5.")
            return redirect(fallback_redirect)

        images = request.FILES.getlist("images")

        if len(images) > 4:
            messages.error(request, "You can only upload a maximum of 4 pictures.")
            return redirect(fallback_redirect)

        review = Review.objects.create(
            user=request.user, booking=booking, rating=rating, comment=comment
        )

        for img in images:
            ReviewImage.objects.create(review=review, image=img)

        log_action(request.user, f"Submitted a review for booking #{booking.id}.")

        # Notify Admin
        AdminNotification.objects.create(
            booking=booking, user=request.user, message="submitted a new review."
        )

        messages.success(request, "Thank you for your review!")
        # On success, if they were in my_reviews, redirect them to my_reviews to see it immediately.
        # Otherwise redirect to the main reviews board.
        if fallback_redirect == "my_reviews":
            return redirect("my_reviews")
        return redirect("reviews")  # Redirect to the new reviews page
    else:
        messages.error(request, "Please provide both a rating and a comment.")

    return redirect(fallback_redirect)


@login_required
def admin_profile(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    return render(request, "admin/admin_profile.html")


# Helper to extract End Time from special_requests string
def get_end_time_from_str(text):
    match = re.search(r"\(End Time: (\d{2}:\d{2})\)", text)
    if match:
        return match.group(1)
    return None


# Helper to format "HH:MM" or time object into readable 12-hour time.
def format_time_12h(value):
    if not value:
        return ""
    try:
        if hasattr(value, "strftime"):
            return value.strftime("%I:%M %p").lstrip("0")
        parsed = datetime.strptime(str(value), "%H:%M")
        return parsed.strftime("%I:%M %p").lstrip("0")
    except (ValueError, TypeError):
        return str(value)


# Helper to remove existing End Time tag to prevent duplication
def remove_end_time_tag(text):
    if not text:
        return ""
    return re.sub(r"\s*\(End Time: \d{2}:\d{2}\)", "", text).strip()


# Helper to format full time range string (e.g., "10:00 AM - 12:00 PM")
def get_booking_time_range(booking):
    if not booking.event_time:
        return ""
    start_str = format_time_12h(booking.event_time)
    end_time_str = get_end_time_from_str(booking.special_requests or "")
    if end_time_str:
        try:
            end_time_obj = datetime.strptime(end_time_str, "%H:%M").time()
            end_str = format_time_12h(end_time_obj)
            return f"{start_str} - {end_str}"
        except ValueError:
            pass
    return start_str


# -------------------------
# BOOKING PAGE (Calendar + Form)
# -------------------------
@login_required
def booking_page(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    design_id = request.GET.get('design_id')
    prefilled_design = None
    if design_id:
        try:
            prefilled_design = UserDesign.objects.select_related('base_package').get(id=design_id, user=request.user)
        except UserDesign.DoesNotExist:
            pass

    # Prepare Calendar Events (Client side: ONLY show approved, confirmed, or completed bookings)
    all_bookings = Booking.objects.filter(
        event_date__gte=timezone.now().date(),
        status__in=["pending_payment", "confirmed", "completed"]
    )
    calendar_events = []

    # Get active packages and optionals for the stepper
    active_packages = Package.objects.filter(is_active=True)
    active_addons = AddOn.objects.filter(is_active=True)
    active_additionals = AdditionalOnly.objects.all()
    service_charge_config = get_service_charge_config()

    for b in all_bookings:
        start_dt = (
            datetime.combine(b.event_date, b.event_time) if b.event_time else None
        )
        end_time_str = get_end_time_from_str(b.special_requests or "")
        end_dt = None
        if end_time_str:
            try:
                end_time_obj = datetime.strptime(end_time_str, "%H:%M").time()
                end_dt = datetime.combine(b.event_date, end_time_obj)
            except ValueError:
                pass

        # Define event color (Client side: All blocked slots appear blue)
        event_color = "#3b82f6"  # Blue for confirmed, completed, and pending_payment

        calendar_events.append(
            {
                "title": get_booking_time_range(b),
                "start": start_dt.isoformat() if start_dt else b.event_date.isoformat(),
                "end": end_dt.isoformat() if end_dt else None,
                "color": event_color,
            }
        )

    return render(
        request,
        "client/booking/booking_page.html",
        {
            "calendar_events": calendar_events,
            "packages": active_packages,
            "active_addons": active_addons,
            "active_additionals": active_additionals,
            "global_service_charge": service_charge_config.amount,
            "global_service_charge_note": service_charge_config.notes,
            "prefilled_design": prefilled_design,
        },
    )


# -------------------------
# ADMIN CALENDAR PAGE
# -------------------------
@login_required
def admin_calendar(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    # Show ALL bookings regardless of status or date
    all_bookings = Booking.objects.select_related("user").all()
    calendar_events = []

    # Color mapping per status
    status_colors = {
        "confirmed": "#3b82f6",  # Blue
        "pending": "#d97706",  # Yellow/Amber
        "completed": "#22c55e",  # Green
        "expired": "#6b7280",  # Gray
        "cancelled": "#ef4444",  # Red
        "cancel_requested": "#f97316",  # Orange
    }

    for b in all_bookings:
        start_dt = (
            datetime.combine(b.event_date, b.event_time) if b.event_time else None
        )
        end_time_str = get_end_time_from_str(b.special_requests or "")
        end_dt = None
        if end_time_str:
            try:
                end_time_obj = datetime.strptime(end_time_str, "%H:%M").time()
                end_dt = datetime.combine(b.event_date, end_time_obj)
            except ValueError:
                pass

        # Include client name and time range in the title
        client_name = b.user.first_name or b.user.username
        time_range = get_booking_time_range(b)
        event_title = f"{client_name} — {time_range}" if time_range else client_name

        # Clean special requests for display
        cleaned_requests = remove_end_time_tag(b.special_requests or "")

        calendar_events.append(
            {
                "title": event_title,
                "start": start_dt.isoformat() if start_dt else b.event_date.isoformat(),
                "end": end_dt.isoformat() if end_dt else None,
                "color": status_colors.get(b.status, "#3b82f6"),
                "booking_id": b.id,
                "client_name": f"{b.user.first_name} {b.user.last_name}".strip()
                or b.user.username,
                "event_type": b.event_type or "—",
                "event_location": b.event_location or "—",
                "package_type": b.package_type or "—",
                "status": b.get_status_display(),
                "status_raw": b.status,
                "time_range": time_range or "—",
                "event_date": b.event_date.strftime("%B %d, %Y"),
                "total_price": str(b.total_price),
            }
        )

    return render(
        request,
        "admin/admin_calendar.html",
        {
            "calendar_events": calendar_events,
        },
    )


@login_required
def mark_notifications_read(request):
    if request.user.role not in ["admin", "staff"]:
        return JsonResponse({"status": "error", "message": "Not allowed"}, status=403)

    if request.method == "POST":
        # Mark all legacy booking notifications as read
        Booking.objects.filter(admin_notified=False).update(admin_notified=True)
        # Mark all new unified events as read
        AdminNotification.objects.filter(is_read=False).update(is_read=True)
        return JsonResponse({"status": "success"})

    return JsonResponse({"status": "error"}, status=400)


@login_required
def hide_notification(request, id):
    if request.user.role not in ["admin", "staff"]:
        return JsonResponse({"status": "error", "message": "Not allowed"}, status=403)

    if request.method == "POST":
        notif_id_str = str(id)
        if notif_id_str.startswith("b_"):
            # It's a legacy booking notification
            real_id = notif_id_str.replace("b_", "")
            booking = get_object_or_404(Booking, id=real_id)
            booking.admin_notif_hidden = True
            booking.save()
        elif notif_id_str.startswith("n_"):
            # It's a new admin notification event
            real_id = notif_id_str.replace("n_", "")
            notif = get_object_or_404(AdminNotification, id=real_id)
            notif.is_hidden = True
            notif.save()
        else:
            # Fallback for old integer IDs that might still exist in cached templates
            try:
                real_id = int(notif_id_str)
                booking = get_object_or_404(Booking, id=real_id)
                booking.admin_notif_hidden = True
                booking.save()
            except ValueError:
                return JsonResponse(
                    {"status": "error", "message": "Invalid ID format"}, status=400
                )

        return JsonResponse({"status": "success"})

    return JsonResponse({"status": "error"}, status=400)


@login_required
def mark_customer_notification_read(request, id):
    """Mark a customer's notification as read via AJAX."""
    try:
        notif = Notification.objects.get(id=id, user=request.user)
        notif.is_read = True
        notif.save()
        return JsonResponse({"status": "success"})
    except Notification.DoesNotExist:
        return JsonResponse(
            {"status": "error", "message": "Notification not found"}, status=404
        )


@login_required
@require_POST
def clear_all_notifications(request):
    """Mark all of a customer's notifications as read."""
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return JsonResponse({"status": "success"})


# -------------------------
# CREATE BOOKING
# -------------------------
@login_required
def create_booking(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    is_ajax = request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest"

    if request.method == "POST":
        start_time = request.POST.get("start_time")
        end_time = request.POST.get("end_time")
        event_date = request.POST.get("event_date")
        special_requests = request.POST.get("special_requests", "")
        special_requests = remove_end_time_tag(special_requests)  # Clean up

        # Combine notes if end_time is provided (since model might only have event_time)
        if end_time:
            special_requests = f"{special_requests}\n(End Time: {end_time})".strip()

        # Validation for past dates/times
        now = timezone.localtime(timezone.now())
        today = now.date()

        if event_date:
            booking_date = datetime.strptime(event_date, "%Y-%m-%d").date()

            # 1. Past Date Check
            if booking_date < today:
                error_msg = "Cannot book a date in the past."
                if is_ajax:
                    return JsonResponse({"success": False, "message": error_msg})
                messages.error(request, error_msg)
                return redirect("booking_page")

            # 2. Past Time Check (If booking is Today)
            if booking_date == today and start_time:
                booking_time = datetime.strptime(start_time, "%H:%M").time()
                if booking_time < now.time():
                    error_msg = "The selected start time has already passed. Please choose a future time."
                    if is_ajax:
                        return JsonResponse({"success": False, "message": error_msg})
                    messages.error(request, error_msg)
                    return redirect("booking_page")

            # 3. Minimum Duration Check (2 hours)
            if start_time and end_time:
                booking_start_dt = datetime.combine(
                    today, datetime.strptime(start_time, "%H:%M").time()
                )
                booking_end_dt = datetime.combine(
                    today, datetime.strptime(end_time, "%H:%M").time()
                )

                # Handle cases where end time is crossing midnight (though forms usually restrict this)
                if booking_end_dt <= booking_start_dt:
                    booking_end_dt += timedelta(days=1)

                duration = booking_end_dt - booking_start_dt
                if duration < timedelta(hours=2):
                    error_msg = "Please choose an end time that is at least 2 hours after the start time."
                    if is_ajax:
                        return JsonResponse({"success": False, "message": error_msg})
                    messages.error(request, error_msg)
                    return redirect("booking_page")

            # 3A. Validate Time Between 7:00 AM and 6:00 PM
            if start_time and end_time:
                start_dt = datetime.strptime(start_time, "%H:%M")
                end_dt = datetime.strptime(end_time, "%H:%M")
                
                start_hour = start_dt.hour
                start_minute = start_dt.minute
                end_hour = end_dt.hour
                end_minute = end_dt.minute
                
                if start_hour < 7 or start_hour >= 18:
                    error_msg = "Booking hours are 7:00 AM to 6:00 PM only. Please choose a start time within this range."
                    if is_ajax:
                        return JsonResponse({"success": False, "message": error_msg})
                    messages.error(request, error_msg)
                    return redirect("booking_page")
            
                if end_hour < 7 or end_hour > 18 or (end_hour == 18 and end_minute != 0):
                    error_msg = "Booking hours are 7:00 AM to 6:00 PM only. End time must be between 7:00 AM and 6:00 PM."
                    if is_ajax:
                        return JsonResponse({"success": False, "message": error_msg})
                    messages.error(request, error_msg)
                    return redirect("booking_page")

                # Check for midnight crossing / reversed range after hours validation
                crosses_midnight = (end_hour < start_hour) or (end_hour == start_hour and end_minute <= start_minute)
                if crosses_midnight:
                    error_msg = "End time must be later than start time on the same day."
                    if is_ajax:
                        return JsonResponse({"success": False, "message": error_msg})
                    messages.error(request, error_msg)
                    return redirect("booking_page")

            # 4. Double Booking / Overlap Check
            if start_time and end_time:
                new_start = datetime.combine(
                    booking_date, datetime.strptime(start_time, "%H:%M").time()
                )
                new_end = datetime.combine(
                    booking_date, datetime.strptime(end_time, "%H:%M").time()
                )

                # Get active bookings for this date (Approved, confirmed, or completed bookings block new bookings)
                existing_bookings = Booking.objects.filter(
                    event_date=booking_date,
                    status__in=["pending_payment", "confirmed", "completed"]
                )

                for b in existing_bookings:
                    if not b.event_time:
                        continue

                    b_start = datetime.combine(booking_date, b.event_time)
                    # Extract end time from stored string or default to +4 hours
                    b_end_str = get_end_time_from_str(b.special_requests)
                    if b_end_str:
                        b_end = datetime.combine(
                            booking_date, datetime.strptime(b_end_str, "%H:%M").time()
                        )
                    else:
                        b_end = b_start + timedelta(
                            hours=4
                        )  # Default duration assumption

                    # Check for Overlap: (StartA < EndB) and (EndA > StartB)
                    if new_start < b_end and new_end > b_start:
                        existing_start = format_time_12h(b.event_time)
                        existing_end = format_time_12h(b_end.time())
                        error_msg = (
                            f"Time slot unavailable. Another booking is already scheduled from "
                            f"{existing_start} to {existing_end}. Please choose a different time."
                        )
                        if is_ajax:
                            return JsonResponse(
                                {"success": False, "message": error_msg}
                            )
                        messages.error(request, error_msg)
                        return redirect("booking_page")

        # 4. Validate total_price
        try:
            total_price_val = Decimal(request.POST.get("total_price", "0"))
        except InvalidOperation:
            error_msg = "Invalid price format."
            if is_ajax:
                return JsonResponse({"success": False, "message": error_msg})
            messages.error(request, error_msg)
            return redirect("booking_page")
        if total_price_val <= 0:
            error_msg = "Total price must be greater than 0."
            if is_ajax:
                return JsonResponse({"success": False, "message": error_msg})
            messages.error(request, error_msg)
            return redirect("booking_page")
        MAX_PRICE = Decimal("99999999.99")
        if total_price_val > MAX_PRICE:
            error_msg = "Total price exceeds the maximum allowed value."
            if is_ajax:
                return JsonResponse({"success": False, "message": error_msg})
            messages.error(request, error_msg)
            return redirect("booking_page")

        booking = Booking.objects.create(
            user=request.user,
            event_type=request.POST.get("event_type"),
            event_date=event_date,
            event_time=start_time,  # Save start time to event_time field
            event_location=request.POST.get("event_location"),
            package_type=request.POST.get("package_type"),
            special_requests=special_requests,
            reference_image=request.FILES.get("reference_image"),
            total_price=total_price_val,
        )

        # Associate with UserDesign if provided
        user_design_id = request.POST.get("user_design_id")
        if user_design_id:
            try:
                user_design = UserDesign.objects.get(id=user_design_id, user=request.user)
                # If you want to create a permanent Design record for this booking:
                Design.objects.create(
                    booking=booking,
                    style=user_design.name,
                    color_palette="Custom",
                    image=user_design.thumbnail, # Use the thumbnail as the design image
                    price_estimate=total_price_val,
                    status='finalized'
                )
            except UserDesign.DoesNotExist:
                pass

        # Save multiple images (up to 4)
        images = request.FILES.getlist("reference_images")

        # Fallback if the form only sent single 'reference_image'
        if not images and request.FILES.get("reference_image"):
            images = [request.FILES.get("reference_image")]

        for img in images[:4]:
            BookingImage.objects.create(booking=booking, image=img)

        log_action(request.user, f"Created a new booking #{booking.id}.")

        if is_ajax:
            # Build event data so frontend can add to calendar dynamically
            event_title = (
                f"{start_time} - {end_time} | {request.POST.get('event_type', '')}"
            )
            event_start = f"{event_date}T{start_time}:00" if start_time else event_date
            event_end = f"{event_date}T{end_time}:00" if end_time else None
            return JsonResponse(
                {
                    "success": True,
                    "message": "Your booking has been successfully submitted! Please wait for admin confirmation.",
                }
            )
        messages.success(request, "Booking created successfully!")
        return redirect("booking_page")

    return redirect("booking_page")


# -------------------------
# VIEW BOOKING
# -------------------------
@login_required
def view_booking(request, id):
    booking = get_object_or_404(Booking, id=id)

    if request.user != booking.user:
        return HttpResponseForbidden("Not allowed")

    booking.time_range_display = get_booking_time_range(booking)
    payment_history = booking.payments.select_related("verified_by").order_by(
        "-created_at"
    )
    total_verified_paid = payment_history.filter(payment_status="verified").aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")
    remaining_balance = (booking.total_price or Decimal("0.00")) - total_verified_paid
    cleaned_requests = remove_end_time_tag(booking.special_requests or "")
    source = (request.GET.get("from") or "").strip().lower()
    if source == "my_payments":
        back_url = reverse("my_payments")
        back_label = "Back to My Payments"
    else:
        back_url = reverse("customer_profile")
        back_label = "Back to Dashboard"

    return render(
        request,
        "client/booking/booking_detail.html",
        {
            "booking": booking,
            "payment_history": payment_history,
            "total_verified_paid": total_verified_paid,
            "remaining_balance": remaining_balance,
            "cleaned_requests": cleaned_requests,
            "back_url": back_url,
            "back_label": back_label,
        },
    )


# -------------------------
# EDIT BOOKING
# -------------------------
@login_required
def edit_booking(request, id):
    booking = get_object_or_404(Booking, id=id)

    if request.user != booking.user:
        return HttpResponseForbidden("Not allowed")

    # Confirmed bookings are final. Only pending bookings can still be edited.
    if booking.status != "pending":
        messages.error(
            request,
            "Confirmed bookings can no longer be edited. Please review details before submission.",
        )
        return redirect("customer_profile")

    if request.method == "POST":
        start_time = request.POST.get("start_time")
        end_time = request.POST.get("end_time")
        special_requests = request.POST.get("special_requests", "")
        special_requests = remove_end_time_tag(
            special_requests
        )  # Clean up before appending

        # --- Validation Logic (Same as Create) ---
        now = timezone.localtime(timezone.now())
        today = now.date()
        booking_date = datetime.strptime(
            request.POST.get("event_date"), "%Y-%m-%d"
        ).date()

        if booking_date < today:
            messages.error(request, "Cannot change to a past date.")
            return redirect("edit_booking", id=id)

        if start_time and end_time:
            new_start = datetime.combine(
                booking_date, datetime.strptime(start_time, "%H:%M").time()
            )
            new_end = datetime.combine(
                booking_date, datetime.strptime(end_time, "%H:%M").time()
            )

            # Handle cases where end time is crossing midnight
            booking_end_dt_calc = new_end
            if booking_end_dt_calc <= new_start:
                booking_end_dt_calc += timedelta(days=1)

            if (booking_end_dt_calc - new_start) < timedelta(hours=2):
                messages.error(
                    request,
                    "Please choose an end time that is at least 2 hours after the start time.",
                )
                return redirect("edit_booking", id=id)

            existing_bookings = (
                Booking.objects.filter(event_date=booking_date)
                .exclude(id=booking.id)
                .filter(status__in=["pending_payment", "confirmed", "completed"])
            )
            for b in existing_bookings:
                b_start = datetime.combine(booking_date, b.event_time)
                b_end_str = get_end_time_from_str(b.special_requests)
                b_end = (
                    datetime.combine(
                        booking_date, datetime.strptime(b_end_str, "%H:%M").time()
                    )
                    if b_end_str
                    else b_start + timedelta(hours=4)
                )

                if new_start < b_end and new_end > b_start:
                    existing_start = format_time_12h(b_start.time())
                    existing_end = format_time_12h(b_end.time())
                    messages.error(
                        request,
                        f"Time slot unavailable. Another booking is already scheduled from "
                        f"{existing_start} to {existing_end}. Please choose a different time.",
                    )
                    return redirect("edit_booking", id=id)
        # -----------------------------------------

        if end_time:
            special_requests = f"{special_requests}\n(End Time: {end_time})".strip()

        booking.event_type = request.POST.get("event_type")
        booking.event_date = request.POST.get("event_date")
        booking.event_time = start_time
        booking.event_location = request.POST.get("event_location")
        booking.package_type = request.POST.get("package_type")
        booking.special_requests = special_requests
        booking.total_price = request.POST.get("total_price")

        if request.FILES.get("reference_image"):
            booking.reference_image = request.FILES.get("reference_image")

        # Remove deleted images
        remove_images = request.POST.getlist("remove_images[]")
        if remove_images:
            if "legacy" in remove_images:
                if booking.reference_image:
                    booking.reference_image.delete(save=False)
                remove_images.remove("legacy")
            if remove_images:
                try:
                    remove_ids = [int(i) for i in remove_images if i.isdigit()]
                    if remove_ids:
                        BookingImage.objects.filter(
                            id__in=remove_ids, booking=booking
                        ).delete()
                except ValueError:
                    pass

        # Add new images
        new_images = request.FILES.getlist("reference_images")
        current_img_count = booking.images.count()
        if booking.reference_image:
            current_img_count += 1

        allowed_new = max(0, 4 - current_img_count)

        for img in new_images[:allowed_new]:
            BookingImage.objects.create(booking=booking, image=img)

        booking.edit_requested = False
        booking.edit_allowed = False
        booking.save()

        log_action(request.user, f"Edited booking #{booking.id}.")
        messages.success(request, "Booking updated successfully!")
        return redirect("customer_profile")

    return render(request, "client/booking/booking_form.html", {"booking": booking})


# -------------------------
# DELETE BOOKING
# -------------------------
@login_required
def delete_booking(request, id):
    booking = get_object_or_404(Booking, id=id)

    if request.user != booking.user:
        return HttpResponseForbidden("Not allowed")

    if booking.status != "pending":
        messages.error(request, "Only pending bookings can be deleted.")
        return redirect("customer_profile")

    if request.method == "POST":
        booking_id = booking.id
        booking.delete()
        log_action(request.user, f"Deleted booking #{booking_id}.")
        messages.success(request, "Booking deleted successfully!")
        return redirect("customer_profile")

    return render(
        request, "client/booking/booking_delete_confirm.html", {"booking": booking}
    )


# -------------------------
# ADMIN APPROVE/DENY BOOKING
# -------------------------
@login_required
def admin_booking_list(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    # Auto-expire pending bookings in the past
    check_booking_expirations()

    status_filter = request.GET.get("status")
    bookings = Booking.objects.order_by("-created_at")

    if status_filter:
        if status_filter == "edit_requested":
            bookings = bookings.filter(edit_requested=True)
        else:
            bookings = bookings.filter(status=status_filter)

    # Search Logic
    search_query = request.GET.get("search")
    if search_query:
        if search_query.startswith("#") and search_query[1:].isdigit():
            # If search query looks like #123, search by ID
            bookings = bookings.filter(id=search_query[1:])
        else:
            bookings = bookings.filter(
                Q(user__username__icontains=search_query)
                | Q(user__email__icontains=search_query)
                | Q(event_type__icontains=search_query)
                | Q(status__icontains=search_query)
                | Q(id__icontains=search_query)
            )

    # Pagination Logic (10 items per page)
    paginator = Paginator(bookings, 10)
    page_number = request.GET.get("page")
    bookings_page = paginator.get_page(page_number)

    # Attach formatted time range for display
    for b in bookings_page:
        b.time_range_display = get_booking_time_range(b)

    return render(
        request,
        "admin/booking/admin_booking_list.html",
        {
            "bookings": bookings_page,
            "search_query": search_query or "",
            "status_filter": status_filter or "",
        },
    )


@login_required
def admin_booking_detail(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    booking = get_object_or_404(Booking, id=id)

    # Add formatted time range
    booking.time_range_display = get_booking_time_range(booking)

    # Clean up special requests for display to hide the end time tag
    cleaned_requests = remove_end_time_tag(booking.special_requests or "")
    price_breakdown = get_booking_price_breakdown(booking)

    return render(
        request,
        "admin/booking/admin_booking_detail.html",
        {
            "booking": booking,
            "cleaned_requests": cleaned_requests,
            "price_breakdown": price_breakdown,
        },
    )


@login_required
def admin_booking_action(request, id, action):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    booking = get_object_or_404(Booking, id=id)

    if action == "approve":
        if booking.status != "pending":
            messages.error(request, "Only pending bookings can be approved.")
            return redirect("admin_booking_list")

        booking.status = "pending_payment"
        booking.admin_denial_reason = None
        booking.save()

        # Auto-cancel other PENDING bookings on the same date that OVERLAP in time
        potential_conflicts = Booking.objects.filter(
            event_date=booking.event_date,
            status="pending",
        ).exclude(id=booking.id)

        # Approved booking time range
        b_start = datetime.combine(booking.event_date, booking.event_time) if booking.event_time else None
        b_end_str = get_end_time_from_str(booking.special_requests)
        if b_end_str:
            b_end = datetime.combine(booking.event_date, datetime.strptime(b_end_str, "%H:%M").time())
        else:
            b_end = b_start + timedelta(hours=4) if b_start else None

        cancelled_count = 0
        for conflict in potential_conflicts:
            # Check for overlap if both have times
            if b_start and b_end and conflict.event_time:
                c_start = datetime.combine(conflict.event_date, conflict.event_time)
                c_end_str = get_end_time_from_str(conflict.special_requests)
                if c_end_str:
                    c_end = datetime.combine(conflict.event_date, datetime.strptime(c_end_str, "%H:%M").time())
                else:
                    c_end = c_start + timedelta(hours=4)

                # Overlap: (StartA < EndB) and (EndA > StartB)
                if b_start < c_end and b_end > c_start:
                    conflict.status = "cancelled"
                    conflict.admin_denial_reason = (
                        "Another booking was approved for this time slot. "
                        "Please choose a different time or date."
                    )
                    conflict.save()
                    Notification.objects.create(
                        user=conflict.user,
                        booking=conflict,
                        message=(
                            f"We're sorry, but your booking #{conflict.id} on "
                            f"{conflict.event_date} was not approved because another "
                            f"booking was already confirmed for that time slot. "
                            f"Please book a different time or date."
                        ),
                    )
                    cancelled_count += 1
            else:
                # If either doesn't have time, we assume they conflict (old behavior for safety)
                # or you might want to allow it. Given the system seems to require times,
                # we'll keep the safe approach of cancelling if date matches and time is missing.
                conflict.status = "cancelled"
                conflict.admin_denial_reason = (
                    "Another booking was approved for this date. "
                    "Please choose a different date."
                )
                conflict.save()
                Notification.objects.create(
                    user=conflict.user,
                    booking=conflict,
                    message=(
                        f"We're sorry, but your booking #{conflict.id} on "
                        f"{conflict.event_date} was not approved because another "
                        f"booking was already confirmed for that date. "
                        f"Please book a different date."
                    ),
                )
                cancelled_count += 1

        # Notify the approved customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=(
                f"Great news! Your booking #{booking.id} has been approved! 🎉 "
                f"Please proceed to payment to confirm your slot."
            ),
        )

        # Admin notification
        AdminNotification.objects.create(
            booking=booking,
            user=booking.user,
            message=(
                f"Booking #{booking.id} approved by {request.user.get_full_name() or request.user.username}."
                + (
                    f" {cancelled_count} conflicting booking(s) auto-cancelled."
                    if cancelled_count
                    else ""
                )
            ),
        )

        log_action(
            request.user,
            f"Approved booking #{booking.id} for '{booking.user.username}'."
            + (
                f" Auto-cancelled {cancelled_count} conflicting booking(s)."
                if cancelled_count
                else ""
            ),
        )
        messages.success(
            request,
            f"Booking #{booking.id} approved! Customer has been notified to proceed with payment."
            + (
                f" {cancelled_count} conflicting booking(s) were auto-cancelled."
                if cancelled_count
                else ""
            ),
        )

    elif action == "confirm":
        booking.status = "confirmed"
        booking.admin_denial_reason = None
        log_action(
            request.user,
            f"Confirmed booking #{booking.id} for '{booking.user.username}'.",
        )
        booking.save()

        # Notify Customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=f"Hooray! Your booking #{booking.id} is now CONFIRMED! See you soon! 🎉",
        )
        messages.success(request, "Booking confirmed!")
    elif action == "deny":
        if request.method != "POST":
            messages.error(request, "Please provide a denial reason.")
            return redirect("admin_booking_list")

        deny_reason = request.POST.get("deny_reason", "").strip()
        if not deny_reason:
            messages.error(request, "Denial reason is required.")
            return redirect("admin_booking_list")

        booking.status = "cancelled"
        booking.admin_denial_reason = deny_reason
        booking.save()

        # Notify Customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=f"We're sorry, but your booking #{booking.id} was NOT approved. Reason: {deny_reason}",
        )
        log_action(
            request.user, f"Denied booking #{booking.id} for '{booking.user.username}'."
        )
        messages.success(request, "Booking denied!")
    elif action == "complete":
        booking.status = "completed"
        booking.admin_denial_reason = None
        booking.save()

        # Notify Customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=f"Thank you for trusting us! Your booking #{booking.id} is now COMPLETED. We hope you enjoyed our service! ❤️ Please feel free to leave a review about your experience! 😊",
        )
        log_action(
            request.user,
            f"Marked booking #{booking.id} as completed for '{booking.user.username}'.",
        )
        messages.success(request, "Booking marked as completed!")

    return redirect("admin_booking_list")


@login_required
def request_cancel_booking(request, id):
    get_object_or_404(Booking, id=id, user=request.user)

    messages.error(
        request,
        "Request cancel is no longer available. Pending bookings can only be canceled before confirmation.",
    )
    return redirect("customer_profile")


@login_required
def admin_cancel_action(request, id, action):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    booking = get_object_or_404(Booking, id=id)

    if booking.status != "cancel_requested":
        messages.error(request, "Invalid action.")
        return redirect("admin_booking_list")

    if action == "approve":
        booking.status = "cancelled"
        booking.cancel_request_reason = None
        log_action(
            request.user, f"Approved cancellation request for booking #{booking.id}."
        )

        # Notify Customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=f"Your cancellation request for booking #{booking.id} has been APPROVED. Your booking is now cancelled.",
        )
        messages.success(request, "Cancellation approved. Booking is now cancelled.")
    elif action == "deny":
        booking.status = "confirmed"
        booking.cancel_request_reason = None
        log_action(
            request.user, f"Denied cancellation request for booking #{booking.id}."
        )

        # Notify Customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=f"We're sorry, but your cancellation request for booking #{booking.id} was NOT approved.",
        )
        messages.success(request, "Cancellation denied. Booking remains confirmed.")

    booking.save()
    return redirect("admin_booking_list")


@login_required
def request_edit_booking(request, id):
    get_object_or_404(Booking, id=id, user=request.user)

    messages.error(
        request,
        "Request edit is no longer available. Confirmed bookings are final and cannot be edited.",
    )
    return redirect("customer_profile")


@login_required
def admin_edit_action(request, id, action):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    booking = get_object_or_404(Booking, id=id)

    if not booking.edit_requested:
        messages.error(request, "No edit request found.")
        return redirect("admin_booking_list")

    if action == "approve":
        booking.edit_requested = False
        booking.edit_allowed = False
        booking.edit_request_reason = None
        booking.edit_original_snapshot = None
        log_action(request.user, f"Approved edit request for booking #{booking.id}.")

        # Notify Customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=f"Your edit request for booking #{booking.id} has been APPROVED. Your updated booking details are now confirmed.",
        )
        messages.success(
            request, "Edit approved. Updated booking details are now confirmed."
        )
    elif action == "deny":
        apply_booking_snapshot(booking, booking.edit_original_snapshot)
        booking.edit_requested = False
        booking.edit_allowed = False
        booking.edit_request_reason = None
        booking.edit_original_snapshot = None
        log_action(request.user, f"Denied edit request for booking #{booking.id}.")

        # Notify Customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=f"We're sorry, but your edit request for booking #{booking.id} was NOT approved.",
        )
        messages.success(request, "Edit denied. Booking reverted to previous details.")

    booking.save()
    return redirect("admin_booking_list")


# =========================
# ADMIN USER MANAGEMENT
# =========================


@login_required
def admin_user_list(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    # 1. Kunin lahat ng users
    users_list = User.objects.all().order_by("username")

    # 2. Role Filter
    role_filter = request.GET.get("role")
    if role_filter and role_filter in ["admin", "staff", "customer"]:
        users_list = users_list.filter(role=role_filter)

    # 3. Search Logic
    search_query = request.GET.get("search")
    if search_query:
        users_list = users_list.filter(
            Q(username__icontains=search_query)
            | Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(email__icontains=search_query)
        )

    # 4. Pagination Logic (10 items per page)
    paginator = Paginator(users_list, 10)
    page_number = request.GET.get("page")
    users = paginator.get_page(page_number)

    return render(request, "admin/user/admin_user_list.html", {"users": users})


@login_required
def admin_user_edit(request, id):
    if request.user.role != "admin":
        return HttpResponseForbidden("Admins only")

    user_obj = get_object_or_404(User, id=id)

    # ❌ bawal i-edit ang sarili
    if user_obj == request.user:
        messages.error(request, "You cannot edit your own account.")
        return redirect("admin_user_list")

    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        email = request.POST.get("email", "").strip()
        phone_number = request.POST.get("phone_number", "").strip()
        role = request.POST.get("role", "").strip()

        user_obj.first_name = first_name
        user_obj.last_name = last_name
        user_obj.email = email
        user_obj.phone_number = phone_number
        user_obj.role = role

        if not first_name:
            messages.error(request, "First name is required.")
            return render(request, "admin/user/admin_user_edit.html", {"u": user_obj})

        if not last_name:
            messages.error(request, "Last name is required.")
            return render(request, "admin/user/admin_user_edit.html", {"u": user_obj})

        if not email:
            messages.error(request, "Email is required.")
            return render(request, "admin/user/admin_user_edit.html", {"u": user_obj})

        if not phone_number:
            messages.error(request, "Phone number is required.")
            return render(request, "admin/user/admin_user_edit.html", {"u": user_obj})

        valid_roles = {"admin", "staff", "customer"}
        if role not in valid_roles:
            messages.error(request, "Role is required.")
            return render(request, "admin/user/admin_user_edit.html", {"u": user_obj})

        log_action(request.user, f"Edited user profile for '{user_obj.username}'.")
        user_obj.save()
        messages.success(request, "User updated successfully.")
        return redirect("admin_user_list")

    return render(request, "admin/user/admin_user_edit.html", {"u": user_obj})


@login_required
def admin_user_toggle_active(request, id):
    if request.user.role != "admin":
        return HttpResponseForbidden("Admins only")

    user_obj = get_object_or_404(User, id=id)

    if user_obj == request.user:
        messages.error(request, "You cannot deactivate yourself.")
        return redirect("admin_user_list")

    user_obj.is_active = not user_obj.is_active
    user_obj.save()

    status = "activated" if user_obj.is_active else "deactivated"
    log_action(
        request.user, f"User account for '{user_obj.username}' has been {status}."
    )
    messages.success(
        request, f"User '{user_obj.username}' has been {status} successfully."
    )
    return redirect("admin_user_list")


@login_required
def admin_user_delete(request, id):
    if request.user.role != "admin":
        return HttpResponseForbidden("Admins only")

    user_obj = get_object_or_404(User, id=id)

    if user_obj == request.user:
        messages.error(request, "You cannot delete yourself.")
        return redirect("admin_user_list")

    username = user_obj.username
    user_obj.delete()
    log_action(request.user, f"Deleted user account for '{username}'.")
    messages.success(request, "User deleted.")
    return redirect("admin_user_list")


@login_required
def admin_about_content(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    content = AboutContent.objects.first()
    if content is None:
        content = AboutContent.objects.create()

    if request.method == "POST":
        content.hero_title = request.POST.get("hero_title", content.hero_title).strip()
        content.hero_subtitle = request.POST.get("hero_subtitle", content.hero_subtitle).strip()
        
        content.story_label = request.POST.get("story_label", content.story_label).strip()
        content.story_title = request.POST.get("story_title", content.story_title).strip()
        content.story_paragraph_1 = request.POST.get("story_paragraph_1", content.story_paragraph_1).strip()
        content.story_paragraph_2 = request.POST.get("story_paragraph_2", content.story_paragraph_2).strip()
        
        content.stat_events_styled = request.POST.get("stat_events_styled", content.stat_events_styled).strip()
        content.stat_year_founded = request.POST.get("stat_year_founded", content.stat_year_founded).strip()
        content.stat_satisfaction = request.POST.get("stat_satisfaction", content.stat_satisfaction).strip()
        
        content.mission_label = request.POST.get("mission_label", content.mission_label).strip()
        content.mission_title = request.POST.get("mission_title", content.mission_title).strip()
        content.mission_paragraph_1 = request.POST.get("mission_paragraph_1", content.mission_paragraph_1).strip()
        content.mission_paragraph_2 = request.POST.get("mission_paragraph_2", content.mission_paragraph_2).strip()
        
        content.values_title = request.POST.get("values_title", content.values_title).strip()
        content.values_subtitle = request.POST.get("values_subtitle", content.values_subtitle).strip()

        if request.FILES.get("story_image"):
            content.story_image = request.FILES["story_image"]
        if request.FILES.get("mission_image"):
            content.mission_image = request.FILES["mission_image"]

        content.save()
        log_action(request.user, "Updated About page content.")
        messages.success(request, "About content updated successfully.")
        return redirect("admin_about_content")

    values = AboutValueItem.objects.all()
    return render(
        request,
        "admin/content/about_content.html",
        {
            "content": content,
            "values": values,
        },
    )


@login_required
def admin_about_value_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    content = AboutContent.objects.first()
    if content is None:
        content = AboutContent.objects.create()

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        description = request.POST.get("description", "").strip()
        icon_class = request.POST.get("icon_class", "fas fa-star").strip()
        is_active = request.POST.get("is_active") == "on"

        try:
            display_order = int(request.POST.get("display_order", 0))
        except (ValueError, TypeError):
            display_order = 0

        if not title:
            messages.error(request, "Title is required.")
            return render(
                request,
                "admin/content/about_value_form.html",
                {
                    "action": "Create",
                    "feature": {},
                    "post_data": request.POST,
                },
            )

        AboutValueItem.objects.create(
            about_content=content,
            title=title,
            description=description,
            icon_class=icon_class,
            display_order=display_order,
            is_active=is_active,
        )
        log_action(request.user, f"Created about value item '{title}'.")
        messages.success(request, "Value item created successfully.")
        return redirect("admin_about_content")

    return render(request, "admin/content/about_value_form.html", {"action": "Create", "feature": {}, "post_data": {}})


@login_required
def admin_about_value_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    value = get_object_or_404(AboutValueItem, id=id)

    if request.method == "POST":
        value.title = request.POST.get("title", value.title).strip()
        value.description = request.POST.get("description", value.description).strip()
        value.icon_class = request.POST.get("icon_class", value.icon_class).strip()
        value.is_active = request.POST.get("is_active") == "on"

        try:
            value.display_order = int(request.POST.get("display_order", value.display_order))
        except (ValueError, TypeError):
            pass

        value.save()
        log_action(request.user, f"Updated about value item '{value.title}' (ID #{value.id}).")
        messages.success(request, "Value item updated successfully.")
        return redirect("admin_about_content")

    return render(
        request,
        "admin/content/about_value_form.html",
        {
            "action": "Edit",
            "feature": value,
            "post_data": {},
        },
    )


@login_required
def admin_about_value_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    value = get_object_or_404(AboutValueItem, id=id)
    value_title = value.title
    value.delete()
    log_action(request.user, f"Deleted about value item '{value_title}' (ID #{id}).")
    messages.success(request, "Value item deleted successfully.")
    return redirect("admin_about_content")


@login_required
def admin_audit_log_list(request):
    if request.user.role not in ["admin"]:
        return HttpResponseForbidden("Admins only")

    log_list = AuditLog.objects.select_related("user").all()

    # Search Logic
    search_query = request.GET.get("search")
    if search_query:
        log_list = log_list.filter(
            Q(action__icontains=search_query)
            | Q(user__username__icontains=search_query)
        )

    # Pagination Logic (15 items per page)
    paginator = Paginator(log_list, 15)
    page_number = request.GET.get("page")
    logs = paginator.get_page(page_number)

    return render(
        request,
        "admin/audit_log_list.html",
        {"logs": logs, "search_query": search_query or ""},
    )


@login_required
def admin_package_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    if request.method == "POST":
        features = request.POST.get("features", "").strip()
        if not features:
            messages.error(request, "Inclusions are required.")
            return render(request, "admin/package/package_form.html")

        try:
            price = Decimal(request.POST["price"])
        except InvalidOperation:
            messages.error(request, "Invalid price format.")
            return render(request, "admin/package/package_form.html")

        MAX_PRICE = Decimal("99999999.99")
        if price < 0 or price > MAX_PRICE:
            messages.error(request, "Price must be between 0 and 99,999,999.99.")
            return render(request, "admin/package/package_form.html")

        package = Package.objects.create(
            name=request.POST.get("name"),
            image=request.FILES.get("image"),
            features=features,
            price=price,
            notes=request.POST.get("notes"),
            is_featured=bool(request.POST.get("is_featured")),
        )

        log_action(request.user, f"Created new package: '{package.name}'.")
        messages.success(request, "Package created successfully!")
        return redirect("admin_package_list")

    return render(request, "admin/package/package_form.html")


@login_required
def admin_package_edit(request, id):
    package = get_object_or_404(Package, id=id)

    if request.method == "POST":
        features = request.POST.get("features", "").strip()
        if not features:
            messages.error(request, "Inclusions are required.")
            return render(
                request, "admin/package/package_form.html", {"package": package}
            )

        try:
            package.price = Decimal(request.POST["price"])
        except InvalidOperation:
            messages.error(request, "Invalid price format.")
            return render(
                request, "admin/package/package_form.html", {"package": package}
            )

        MAX_PRICE = Decimal("99999999.99")
        if package.price < 0 or package.price > MAX_PRICE:
            messages.error(request, "Price must be between 0 and 99,999,999.99.")
            return render(
                request, "admin/package/package_form.html", {"package": package}
            )

        package.name = request.POST.get("name")
        package.features = features
        package.notes = request.POST.get("notes")
        package.is_featured = bool(request.POST.get("is_featured"))

        if request.FILES.get("image"):
            package.image = request.FILES.get("image")

        package.save()
        log_action(request.user, f"Edited package: '{package.name}'.")
        messages.success(request, "Package updated successfully!")
        return redirect("admin_package_list")

    return render(request, "admin/package/package_form.html", {"package": package})


@login_required
def admin_package_list(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    # Featured packages first, then others, newest first
    packages = Package.objects.all().order_by("-is_featured", "-created_at")
    addons = AddOn.objects.all().order_by("-created_at")
    additionals = AdditionalOnly.objects.all().order_by("-created_at")
    service_charge_config = get_service_charge_config()

    return render(
        request,
        "admin/package/package_list.html",
        {
            "packages": packages,
            "addons": addons,
            "additionals": additionals,
            "service_charge_config": service_charge_config,
        },
    )


@login_required
def admin_service_charge_update(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    if request.method != "POST":
        return redirect("admin_package_list")

    config = get_service_charge_config()
    amount_raw = request.POST.get("service_charge_amount", "").strip()
    notes = request.POST.get("service_charge_notes", "").strip()

    if not notes:
        messages.error(request, "Service charge notes are required.")
        return redirect("admin_package_list")

    if not amount_raw:
        messages.error(request, "Service charge amount is required.")
        return redirect("admin_package_list")

    try:
        amount = Decimal(amount_raw)
    except InvalidOperation:
        messages.error(request, "Invalid service charge amount.")
        return redirect("admin_package_list")

    max_price = Decimal("99999999.99")
    if amount < 0 or amount > max_price:
        messages.error(request, "Service charge must be between 0 and 99,999,999.99.")
        return redirect("admin_package_list")

    config.amount = amount
    config.notes = notes
    config.save()

    log_action(request.user, f"Updated global service charge to {config.amount}.")
    messages.success(request, "Service charge settings updated.")
    return redirect("admin_package_list")


@login_required
def admin_package_detail(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    package = get_object_or_404(Package, id=id)

    return render(request, "admin/package/package_detail.html", {"package": package})


@login_required
def admin_package_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    package = get_object_or_404(Package, id=id)
    package_name = package.name
    package.delete()
    log_action(request.user, f"Deleted package: '{package_name}'.")
    messages.success(request, "Package deleted successfully!")
    return redirect("admin_package_list")


def package(request):
    packages = Package.objects.all().order_by("-is_featured", "-created_at")
    addons = AddOn.objects.filter(is_active=True).order_by("-created_at")
    additionals = AdditionalOnly.objects.filter(is_active=True).order_by("-created_at")
    service_charge_config = get_service_charge_config()

    return render(
        request,
        "client/package.html",
        {
            "packages": packages,
            "addons": addons,
            "additionals": additionals,
            "service_charge_amount": service_charge_config.amount,
            "service_charge_notes": service_charge_config.notes,
        },
    )


from decimal import Decimal, InvalidOperation


@login_required
def admin_addon_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    if request.method == "POST":
        features = request.POST.get("features", "").strip()
        if not features:
            messages.error(request, "Inclusions are required.")
            return render(request, "admin/package/addon_form.html")

        solo_raw = request.POST.get("solo_price", "").strip()
        if not solo_raw:
            messages.error(request, "Solo price is required.")
            return render(request, "admin/package/addon_form.html")

        try:
            price = Decimal(request.POST["price"])
            solo_price = Decimal(solo_raw)
        except InvalidOperation:
            messages.error(request, "Invalid price format.")
            return render(request, "admin/package/addon_form.html")

        MAX_PRICE = Decimal("99999999.99")
        if price < 0 or price > MAX_PRICE or solo_price < 0 or solo_price > MAX_PRICE:
            messages.error(request, "Price must be between 0 and 99,999,999.99.")
            return render(request, "admin/package/addon_form.html")

        addon = AddOn.objects.create(
            name=request.POST.get("name"),
            image=request.FILES.get("image"),
            price=price,
            solo_price=solo_price,
            features=features,
        )

        log_action(request.user, f"Created new add-on: '{addon.name}'.")
        messages.success(request, "Add-on created successfully!")
        return redirect("admin_package_list")

    return render(request, "admin/package/addon_form.html")


@login_required
def admin_addon_detail(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    addon = get_object_or_404(AddOn, id=id)

    return render(request, "admin/package/addon_detail.html", {"addon": addon})


@login_required
def admin_addon_edit(request, id):
    addon = get_object_or_404(AddOn, id=id)

    if request.method == "POST":
        features = request.POST.get("features", "").strip()
        if not features:
            messages.error(request, "Inclusions are required.")
            return render(request, "admin/package/addon_form.html", {"addon": addon})

        solo_raw = request.POST.get("solo_price", "").strip()
        if not solo_raw:
            messages.error(request, "Solo price is required.")
            return render(request, "admin/package/addon_form.html", {"addon": addon})

        try:
            addon.price = Decimal(request.POST["price"])
            addon.solo_price = Decimal(solo_raw)
        except InvalidOperation:
            messages.error(request, "Invalid price format.")
            return render(request, "admin/package/addon_form.html", {"addon": addon})

        MAX_PRICE = Decimal("99999999.99")
        if (
            addon.price < 0
            or addon.price > MAX_PRICE
            or addon.solo_price < 0
            or addon.solo_price > MAX_PRICE
        ):
            messages.error(request, "Price must be between 0 and 99,999,999.99.")
            return render(request, "admin/package/addon_form.html", {"addon": addon})

        addon.name = request.POST.get("name")
        addon.features = features

        if request.FILES.get("image"):
            addon.image = request.FILES.get("image")

        addon.save()

        log_action(request.user, f"Edited add-on: '{addon.name}'.")
        messages.success(request, "Add-on updated successfully!")
        return redirect("admin_package_list")

    return render(request, "admin/package/addon_form.html", {"addon": addon})


@login_required
def admin_addon_delete(request, id):
    addon = get_object_or_404(AddOn, id=id)
    addon_name = addon.name
    addon.delete()
    log_action(request.user, f"Deleted add-on: '{addon_name}'.")
    messages.success(request, "Add-on deleted successfully!")
    return redirect("admin_package_list")


@login_required
def admin_additional_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    if request.method == "POST":
        features = request.POST.get("features", "").strip()
        if not features:
            messages.error(request, "Inclusions are required.")
            return render(request, "admin/package/additional_form.html")

        try:
            price = Decimal(request.POST["price"])
        except InvalidOperation:
            messages.error(request, "Invalid price format.")
            return render(request, "admin/package/additional_form.html")

        MAX_PRICE = Decimal("99999999.99")
        if price < 0 or price > MAX_PRICE:
            messages.error(request, "Price must be between 0 and 99,999,999.99.")
            return render(request, "admin/package/additional_form.html")

        additional = AdditionalOnly.objects.create(
            name=request.POST.get("name"),
            image=request.FILES.get("image"),
            price=price,
            features=features,
            notes=request.POST.get("notes"),
        )

        log_action(request.user, f"Created new additional item: '{additional.name}'.")
        messages.success(request, "Additional created successfully!")
        return redirect("admin_package_list")

    return render(request, "admin/package/additional_form.html")


@login_required
def admin_additional_edit(request, id):
    additional = get_object_or_404(AdditionalOnly, id=id)

    if request.method == "POST":
        features = request.POST.get("features", "").strip()
        if not features:
            messages.error(request, "Inclusions are required.")
            return render(
                request,
                "admin/package/additional_form.html",
                {"additional": additional},
            )

        try:
            additional.price = Decimal(request.POST["price"])
        except InvalidOperation:
            messages.error(request, "Invalid price format.")
            return render(
                request,
                "admin/package/additional_form.html",
                {"additional": additional},
            )

        MAX_PRICE = Decimal("99999999.99")
        if additional.price < 0 or additional.price > MAX_PRICE:
            messages.error(request, "Price must be between 0 and 99,999,999.99.")
            return render(
                request,
                "admin/package/additional_form.html",
                {"additional": additional},
            )

        additional.name = request.POST.get("name")
        additional.features = features
        additional.notes = request.POST.get("notes")

        if request.FILES.get("image"):
            additional.image = request.FILES.get("image")

        additional.save()

        log_action(request.user, f"Edited additional item: '{additional.name}'.")
        messages.success(request, "Additional updated successfully!")
        return redirect("admin_package_list")

    return render(
        request, "admin/package/additional_form.html", {"additional": additional}
    )


@login_required
def admin_additional_detail(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    additional = get_object_or_404(AdditionalOnly, id=id)

    return render(
        request, "admin/package/additional_detail.html", {"additional": additional}
    )


@login_required
def admin_additional_delete(request, id):
    additional = get_object_or_404(AdditionalOnly, id=id)
    additional_name = additional.name
    additional.delete()
    log_action(request.user, f"Deleted additional item: '{additional_name}'.")
    messages.success(request, "Additional deleted successfully!")
    return redirect("admin_package_list")


# -----------------------------
# 13️⃣ Service Page Admin
# -----------------------------

@login_required
def admin_service_content(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    content = ServiceContent.objects.first()
    if content is None:
        content = ServiceContent.objects.create(
            hero_title="Our Services",
            hero_subtitle="Balloon styling and event decoration services for all types of events."
        )

    if request.method == "POST":
        content.hero_title = request.POST.get("hero_title", content.hero_title).strip()
        content.hero_subtitle = request.POST.get("hero_subtitle", content.hero_subtitle).strip()
        content.save()
        
        log_action(request.user, "Updated Service page content.")
        messages.success(request, "Service content updated successfully.")
        return redirect("admin_service_content")

    services = Service.objects.all()
    return render(
        request,
        "admin/content/service_content.html",
        {
            "content": content,
            "services": services,
        },
    )


@login_required
def admin_service_item_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        description = request.POST.get("description", "").strip()
        is_active = request.POST.get("is_active") == "on"

        try:
            display_order = int(request.POST.get("display_order", 0))
        except (ValueError, TypeError):
            display_order = 0

        if not title:
            messages.error(request, "Title is required.")
            return render(
                request,
                "admin/content/service_item_form.html",
                {
                    "action": "Create",
                    "feature": {},
                    "post_data": request.POST,
                },
            )

        service = Service.objects.create(
            title=title,
            description=description,
            display_order=display_order,
            is_active=is_active,
        )
        
        if request.FILES.get("image"):
            service.image = request.FILES["image"]
            service.save()

        log_action(request.user, f"Created service item '{title}'.")
        messages.success(request, "Service item created successfully.")
        return redirect("admin_service_content")

    return render(request, "admin/content/service_item_form.html", {"action": "Create", "feature": {}, "post_data": {}})


@login_required
def admin_service_item_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    service = get_object_or_404(Service, id=id)

    if request.method == "POST":
        service.title = request.POST.get("title", service.title).strip()
        service.description = request.POST.get("description", service.description).strip()
        service.is_active = request.POST.get("is_active") == "on"

        try:
            service.display_order = int(request.POST.get("display_order", service.display_order))
        except (ValueError, TypeError):
            pass

        if request.FILES.get("image"):
            service.image = request.FILES["image"]

        service.save()
        log_action(request.user, f"Updated service item '{service.title}' (ID #{service.id}).")
        messages.success(request, "Service item updated successfully.")
        return redirect("admin_service_content")

    return render(
        request,
        "admin/content/service_item_form.html",
        {
            "action": "Edit",
            "feature": service,
            "post_data": {},
        },
    )


@login_required
def admin_service_item_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    service = get_object_or_404(Service, id=id)
    service_title = service.title
    service.delete()
    log_action(request.user, f"Deleted service item '{service_title}' (ID #{id}).")
    messages.success(request, "Service item deleted successfully.")
    return redirect("admin_service_content")


# =========================
# ADMIN REPORTS
# =========================
def _get_reporting_date_range(request):
    filter_preset = request.GET.get("filter_preset")
    today = timezone.now().date()
    
    if filter_preset == "today":
        start_date = today
        end_date = today
    elif filter_preset == "weekly":
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif filter_preset == "monthly":
        start_date = today.replace(day=1)
        _, last_day = monthrange(today.year, today.month)
        end_date = today.replace(day=last_day)
    elif filter_preset == "all_time":
        start_date = datetime(2020, 1, 1).date()
        end_date = today
    else:
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")

        if not start_date_str or not end_date_str:
            end_date = today
            start_date = end_date - timedelta(days=30)
        else:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                end_date = today
                start_date = end_date - timedelta(days=30)

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    return start_date, end_date


def _status_distribution_from_queryset(filtered_bookings):
    status_colors_map = {
        "completed": "#10b981",
        "confirmed": "#3b82f6",
        "pending": "#f59e0b",
        "cancel_requested": "#f87171",
        "cancelled": "#ef4444",
        "expired": "#9ca3af",
    }
    status_dist_qs = (
        filtered_bookings.values("status")
        .annotate(count=Count("id"))
        .order_by("status")
    )
    status_distribution = []
    for item in status_dist_qs:
        status_raw = item["status"]
        status_distribution.append(
            {
                "status": status_raw,
                "label": status_raw.title(),
                "count": item["count"],
                "color": status_colors_map.get(status_raw.lower(), "#6b7280"),
            }
        )
    return status_distribution


def _get_trend_bucket_mode(date_range_days):
    if date_range_days <= 31:
        return "day"
    if date_range_days <= 180:
        return "week"
    return "month"


def _build_trend_spans_and_labels(start_date, end_date, bucket_mode):
    spans = []
    labels = []
    cursor = start_date

    while cursor <= end_date:
        if bucket_mode == "day":
            bucket_end = cursor
            label = cursor.strftime("%b %d")
        elif bucket_mode == "week":
            bucket_end = min(cursor + timedelta(days=6), end_date)
            label = f"{cursor.strftime('%b %d')} - {bucket_end.strftime('%b %d')}"
        else:
            month_end_day = monthrange(cursor.year, cursor.month)[1]
            natural_month_end = datetime(
                cursor.year, cursor.month, month_end_day
            ).date()
            bucket_end = min(natural_month_end, end_date)
            if cursor.day == 1 and bucket_end.day == month_end_day:
                label = cursor.strftime("%b %Y")
            else:
                label = f"{cursor.strftime('%b %d')} - {bucket_end.strftime('%b %d')}"

        spans.append((bucket_end - cursor).days + 1)
        labels.append(label)
        cursor = bucket_end + timedelta(days=1)

    return spans, labels


def _aggregate_trend_series(bookings_qs, range_start, bucket_spans):
    bucket_index_by_day = {}
    cursor = range_start
    for idx, span_days in enumerate(bucket_spans):
        for _ in range(span_days):
            bucket_index_by_day[cursor] = idx
            cursor += timedelta(days=1)

    bookings_series = [0] * len(bucket_spans)
    revenue_series = [0.0] * len(bucket_spans)

    for created_date, booking_total, status in bookings_qs.values_list(
        "created_at__date", "total_price", "status"
    ):
        bucket_idx = bucket_index_by_day.get(created_date)
        if bucket_idx is None:
            continue
        bookings_series[bucket_idx] += 1
        if status == "completed":
            revenue_series[bucket_idx] += float(booking_total or 0)

    return bookings_series, revenue_series


def build_dashboard_context(request):
    start_date, end_date = _get_reporting_date_range(request)
    date_range_days = (end_date - start_date).days + 1

    event_type_options = list(
        Booking.objects.exclude(event_type__isnull=True)
        .exclude(event_type__exact="")
        .values_list("event_type", flat=True)
        .distinct()
        .order_by("event_type")
    )
    selected_event_type = (request.GET.get("event_type") or "all").strip()
    if selected_event_type != "all" and selected_event_type not in event_type_options:
        selected_event_type = "all"

    filtered_bookings = Booking.objects.filter(
        created_at__date__gte=start_date, created_at__date__lte=end_date
    )
    if selected_event_type != "all":
        filtered_bookings = filtered_bookings.filter(event_type=selected_event_type)

    previous_end = start_date - timedelta(days=1)
    previous_start = previous_end - timedelta(days=max(date_range_days - 1, 0))
    previous_period_bookings = Booking.objects.filter(
        created_at__date__gte=previous_start, created_at__date__lte=previous_end
    )
    if selected_event_type != "all":
        previous_period_bookings = previous_period_bookings.filter(
            event_type=selected_event_type
        )

    status_distribution = _status_distribution_from_queryset(filtered_bookings)
    today = timezone.localdate()

    edit_requests = filtered_bookings.filter(edit_requested=True).count()
    cancel_requests = filtered_bookings.filter(status="cancel_requested").count()
    pending_approvals = filtered_bookings.filter(status="pending").count()
    action_queue_total = pending_approvals + edit_requests + cancel_requests

    upcoming_deadline_bookings = (
        Booking.objects.filter(
            status="pending",
            event_date__gte=today,
            event_date__lte=today + timedelta(days=3),
        )
        .select_related("user")
        .order_by("event_date")[:6]
    )
    for booking in upcoming_deadline_bookings:
        booking.days_left = (booking.event_date - today).days

    total_revenue = (
        filtered_bookings.filter(status="completed").aggregate(Sum("total_price"))[
            "total_price__sum"
        ]
        or 0
    )
    total_bookings = filtered_bookings.count()
    completed_count = filtered_bookings.filter(status="completed").count()
    cancelled_count = filtered_bookings.filter(status="cancelled").count()
    avg_booking_value = (
        filtered_bookings.filter(status="completed").aggregate(Avg("total_price"))[
            "total_price__avg"
        ]
        or 0
    )
    completion_rate = (
        round((completed_count / total_bookings) * 100, 1) if total_bookings else 0
    )
    cancellation_rate = (
        round((cancelled_count / total_bookings) * 100, 1) if total_bookings else 0
    )

    prev_period_bookings = previous_period_bookings.count()
    period_delta = total_bookings - prev_period_bookings
    period_delta_pct = (
        round((period_delta / prev_period_bookings) * 100, 1)
        if prev_period_bookings > 0
        else (100.0 if total_bookings > 0 else 0.0)
    )
    prev_completed_revenue = (
        previous_period_bookings.filter(status="completed").aggregate(
            Sum("total_price")
        )["total_price__sum"]
        or 0
    )
    revenue_delta = total_revenue - prev_completed_revenue
    revenue_delta_pct = (
        round((revenue_delta / prev_completed_revenue) * 100, 1)
        if prev_completed_revenue
        else (100.0 if total_revenue > 0 else 0.0)
    )

    revenue_by_event = list(
        filtered_bookings.filter(status="completed")
        .values("event_type")
        .annotate(
            count=Count("id"),
            revenue=Sum("total_price"),
            avg_value=Avg("total_price")
        )
        .order_by("-revenue")
    )
    top_event_label = "No completed bookings yet"
    top_event_revenue = 0
    if revenue_by_event:
        top_event_label = revenue_by_event[0]["event_type"]
        top_event_revenue = revenue_by_event[0]["revenue"] or 0

    package_counts = {}
    completed_with_packages = filtered_bookings.filter(status="completed").values_list(
        "package_type", flat=True
    )
    for package_type in completed_with_packages:
        if not package_type:
            continue
        first_part = str(package_type).split("+")[0].strip()
        if first_part:
            package_counts[first_part] = package_counts.get(first_part, 0) + 1
    top_package_name = "No package data"
    top_package_count = 0
    if package_counts:
        top_package_name, top_package_count = max(
            package_counts.items(), key=lambda item: item[1]
        )

    package_rows = []
    package_revenue = {}
    completed_with_package_revenue = filtered_bookings.filter(
        status="completed"
    ).values_list("package_type", "total_price")
    for package_type, booking_total in completed_with_package_revenue:
        if not package_type:
            continue
        package_name = str(package_type).split("+")[0].strip()
        if not package_name:
            continue
        package_revenue[package_name] = package_revenue.get(
            package_name, Decimal("0.00")
        ) + (booking_total or Decimal("0.00"))

    for package_name, count in sorted(
        package_counts.items(), key=lambda item: item[1], reverse=True
    )[:8]:
        package_rows.append(
            {
                "package_name": package_name,
                "count": count,
                "revenue": package_revenue.get(package_name, Decimal("0.00")),
            }
        )

    top_customers = list(
        filtered_bookings.values("user__first_name", "user__last_name", "user__username", "user__email")
        .annotate(booking_count=Count("id"), total_spent=Sum("total_price"))
        .order_by("-total_spent")[:8]
    )
    for customer in top_customers:
        customer['name'] = f"{customer['user__first_name']} {customer['user__last_name']}".strip() or customer['user__username']

    status_table = [
        {
            "label": item["label"],
            "count": item["count"],
            "color": item["color"],
            "share_pct": round((item["count"] / total_bookings) * 100, 1)
            if total_bookings
            else 0,
        }
        for item in status_distribution
    ]

    recent_audit_logs = AuditLog.objects.select_related("user").order_by("-created_at")[
        :6
    ]
    recent_bookings = filtered_bookings.select_related("user").order_by("-created_at")[
        :15
    ]

    trend_bucket_mode = _get_trend_bucket_mode(date_range_days)
    trend_bucket_spans, chart_labels = _build_trend_spans_and_labels(
        start_date, end_date, trend_bucket_mode
    )
    bookings_trend, revenue_trend = _aggregate_trend_series(
        filtered_bookings, start_date, trend_bucket_spans
    )

    prev_bookings_trend, prev_revenue_trend = _aggregate_trend_series(
        previous_period_bookings,
        previous_start,
        trend_bucket_spans,
    )

    trend_title_map = {
        "day": "Daily Trend",
        "week": "Weekly Trend",
        "month": "Monthly Trend",
    }
    trend_bucket_label_map = {
        "day": "Daily buckets",
        "week": "Weekly buckets",
        "month": "Monthly buckets",
    }

    queue_breakdown = {
        "labels": ["Pending", "Edit Requests", "Cancel Requests"],
        "values": [pending_approvals, edit_requests, cancel_requests],
        "colors": ["#f59e0b", "#3b82f6", "#ef4444"],
    }

    # Busiest days of the week
    busiest_days_data = (
        filtered_bookings.exclude(event_date__isnull=True)
        .values("event_date__week_day")
        .annotate(count=Count("id"))
        .order_by("event_date__week_day")
    )
    
    # Map Django's week_day (1=Sunday, 7=Saturday) to labels
    day_map = {1: "Sun", 2: "Mon", 3: "Tue", 4: "Wed", 5: "Thu", 6: "Fri", 7: "Sat"}
    busiest_days_labels = [day_map.get(item["event_date__week_day"], "Unknown") for item in busiest_days_data]
    busiest_days_values = [item["count"] for item in busiest_days_data]

    # Customer Retention (Returning vs New)
    # Define returning as someone who has more than 1 booking ever
    all_time_customer_counts = (
        Booking.objects.values("user")
        .annotate(count=Count("id"))
    )
    returning_user_ids = [item["user"] for item in all_time_customer_counts if item["count"] > 1]
    
    new_customers_count = filtered_bookings.exclude(user_id__in=returning_user_ids).values("user").distinct().count()
    returning_customers_count = filtered_bookings.filter(user_id__in=returning_user_ids).values("user").distinct().count()
    
    total_customers = new_customers_count + returning_customers_count
    new_customers_pct = round((new_customers_count / total_customers * 100), 1) if total_customers > 0 else 0
    returning_customers_pct = round((returning_customers_count / total_customers * 100), 1) if total_customers > 0 else 0

    return {
        "filter_preset": request.GET.get("filter_preset", ""),
        "start_date": start_date.strftime("%Y-%m-%d"),
        "end_date": end_date.strftime("%Y-%m-%d"),
        "selected_event_type": selected_event_type,
        "event_type_options": event_type_options,
        "active_users": User.objects.filter(role="customer", is_active=True).count(),
        "pending_approvals": pending_approvals,
        "edit_requests": edit_requests,
        "cancel_requests": cancel_requests,
        "action_queue_total": action_queue_total,
        "upcoming_deadline_bookings": upcoming_deadline_bookings,
        "period_delta": period_delta,
        "period_delta_pct": period_delta_pct,
        "total_revenue": total_revenue,
        "avg_booking_value": avg_booking_value,
        "completion_rate": completion_rate,
        "cancellation_rate": cancellation_rate,
        "completed_count": completed_count,
        "cancelled_count": cancelled_count,
        "total_bookings": total_bookings,
        "revenue_delta": revenue_delta,
        "revenue_delta_pct": revenue_delta_pct,
        "top_event_label": top_event_label,
        "top_event_revenue": top_event_revenue,
        "top_package_name": top_package_name,
        "top_package_count": top_package_count,
        "top_customers": top_customers,
        "status_table": status_table,
        "package_rows": package_rows,
        "revenue_by_event": revenue_by_event,
        "date_range_days": date_range_days,
        "recent_audit_logs": recent_audit_logs,
        "recent_bookings": recent_bookings,
        "trend_title": trend_title_map.get(trend_bucket_mode, "Trend"),
        "trend_bucket_label": trend_bucket_label_map.get(
            trend_bucket_mode, "Trend buckets"
        ),
        "dashboard_trend_labels": chart_labels,
        "dashboard_bookings_trend": bookings_trend,
        "dashboard_revenue_trend": revenue_trend,
        "dashboard_prev_bookings_trend": prev_bookings_trend,
        "dashboard_prev_revenue_trend": prev_revenue_trend,
        "queue_breakdown": queue_breakdown,
        "dashboard_status_labels": [item["label"] for item in status_distribution],
        "dashboard_status_values": [item["count"] for item in status_distribution],
        "dashboard_status_colors": [item["color"] for item in status_distribution],
        "dashboard_busiest_days_labels": busiest_days_labels,
        "dashboard_busiest_days_values": busiest_days_values,
        "new_customers_count": new_customers_count,
        "returning_customers_count": returning_customers_count,
        "new_customers_pct": new_customers_pct,
        "returning_customers_pct": returning_customers_pct,
    }


def build_concerns_context(request):
    start_date, end_date = _get_reporting_date_range(request)
    concern_base_qs = (
        ConcernTicket.objects.select_related("user")
        .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
        .order_by("-created_at")
    )

    status_filter = (request.GET.get("concern_status") or "").strip()
    valid_statuses = {choice[0] for choice in ConcernTicket.STATUS_CHOICES}
    if status_filter and status_filter in valid_statuses:
        concern_filtered_qs = concern_base_qs.filter(status=status_filter)
    else:
        status_filter = ""
        concern_filtered_qs = concern_base_qs

    recent_admin_notifications = (
        AdminNotification.objects.select_related("user", "booking")
        .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
        .order_by("-created_at")[:20]
    )

    concern_count = concern_base_qs.count()
    new_count = concern_base_qs.filter(status="new").count()
    in_progress_count = concern_base_qs.filter(status="in_progress").count()
    resolved_count = concern_base_qs.filter(status="resolved").count()

    concerns_paginator = Paginator(concern_filtered_qs, 8)
    concerns_page = concerns_paginator.get_page(request.GET.get("page"))

    return {
        "filter_preset": request.GET.get("filter_preset", ""),
        "start_date": start_date.strftime("%Y-%m-%d"),
        "end_date": end_date.strftime("%Y-%m-%d"),
        "date_range_days": (end_date - start_date).days + 1,
        "concern_rows": concerns_page,
        "concerns_page": concerns_page,
        "concern_status_filter": status_filter,
        "concern_records_count": concern_filtered_qs.count(),
        "total_concerns": concern_count,
        "new_concerns": new_count,
        "in_progress_concerns": in_progress_count,
        "resolved_concerns": resolved_count,
        "admin_notifications": recent_admin_notifications,
    }


@login_required
@login_required
@require_POST
def admin_concern_update(request, id):
    if request.user.role != "admin":
        return HttpResponseForbidden("Admins only")

    ticket = get_object_or_404(ConcernTicket, id=id)
    next_url = request.POST.get("next") or reverse("admin_concerns")

    raw_status = (request.POST.get("status") or "").strip()
    valid_statuses = {choice[0] for choice in ConcernTicket.STATUS_CHOICES}
    if raw_status not in valid_statuses:
        messages.error(request, "Invalid concern status selected.")
        return redirect(next_url)

    ticket.status = raw_status
    ticket.admin_notes = (request.POST.get("admin_notes") or "").strip()
    ticket.save(update_fields=["status", "admin_notes", "updated_at"])

    log_action(
        request.user,
        f"Updated concern #{ticket.id} to '{ticket.get_status_display()}'.",
    )
    messages.success(
        request, f"Concern #{ticket.id} updated to {ticket.get_status_display()}."
    )
    return redirect(next_url)


@require_POST
def chat_api(request):
    """
    API endpoint for the chatbot.
    Expects JSON: { "message": "user message", "session_id": 123 }
    """
    try:
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON body"}, status=400)

        user_message = data.get("message")
        session_id = data.get("session_id")

        if not user_message:
            return JsonResponse({"error": "Message is required"}, status=400)

        # Ensure user is authenticated to use sessions properly
        if not request.user.is_authenticated:
            return JsonResponse({"error": "Authentication required"}, status=403)

        # Handle ChatSession
        session = None
        is_new_session = False
        if session_id:
            try:
                session = ChatSession.objects.get(id=session_id, user=request.user)
            except ChatSession.DoesNotExist:
                pass  # If passed session_id is invalid, we will just create a new one

        # --- 1. FETCH CONTEXT (HISTORY) FIRST ---
        history = []
        if session:
            recent_msgs = ChatMessage.objects.filter(session=session).order_by("-sent_at")[
                :8
            ]

            # Reorder to chronological (oldest to newest) for the AI
            for msg in reversed(recent_msgs):
                role = "user" if msg.sender == request.user else "assistant"
                history.append({"role": role, "content": msg.message})

        # Get AI Response with History
        ai_result = get_chatbot_response(
            user_message,
            conversation_history=history,
            user=request.user,
        )

        # ai_result includes moderation metadata and response text.
        ai_text = ai_result.get("text", "")
        is_warning = ai_result.get("is_warning", False)
        is_banned = ai_result.get("is_banned", False)
        ban_remaining_seconds = int(ai_result.get("ban_remaining_seconds") or 0)
        moderation_action = ai_result.get("moderation_action", "")
        strike_count = int(ai_result.get("strike_count") or 0)
        should_save = ai_result.get("should_save", not is_warning)

        if should_save:
            # Default receiver for AI bot
            admin_user = (
                User.objects.filter(role="admin").first()
                or User.objects.filter(is_superuser=True).first()
            )
            if not admin_user:
                return JsonResponse(
                    {"error": "System misconfiguration (no admin found)"}, status=500
                )

            if not session:
                title = (
                    user_message[:30] + "..." if len(user_message) > 30 else user_message
                )
                session = ChatSession.objects.create(user=request.user, title=title)
                is_new_session = True

            ChatMessage.objects.create(
                session=session,
                sender=request.user,
                receiver=admin_user,
                message=user_message,
                is_flagged=is_warning,
            )

            ChatMessage.objects.create(
                session=session,
                sender=admin_user,
                receiver=request.user,
                message=ai_text,
            )
        elif ai_text and session:
            # Persist moderation/system assistant messages in the active session so
            # chat history remains consistent after reload.
            admin_user = (
                User.objects.filter(role="admin").first()
                or User.objects.filter(is_superuser=True).first()
            )
            if admin_user:
                ChatMessage.objects.create(
                    session=session,
                    sender=admin_user,
                    receiver=request.user,
                    message=ai_text,
                    is_flagged=True,
                )

        return JsonResponse(
            {
                "response": ai_text,
                "is_warning": is_warning,
                "is_banned": is_banned,
                "ban_remaining_seconds": ban_remaining_seconds,
                "moderation_action": moderation_action,
                "strike_count": strike_count,
                "session_id": session.id if session else None,
                "is_new_session": is_new_session,
                "session_title": session.title if session else "",
            }
        )

    except Exception as e:
        print(f"Chat API Error: {e}")  # Debugging
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@require_GET
def chat_sessions(request):
    """
    GET endpoint to fetch all chat sessions for the current user.
    """
    sessions = ChatSession.objects.filter(user=request.user).order_by("-updated_at")
    sessions_list = []
    for s in sessions:
        sessions_list.append(
            {
                "id": s.id,
                "title": s.title,
                "updated_at": s.updated_at.strftime("%b %d, %Y"),
            }
        )
    ban_status = get_current_ban_status(request.user)
    return JsonResponse(
        {
            "sessions": sessions_list,
            "is_banned": ban_status["is_banned"],
            "ban_remaining_seconds": ban_status["ban_remaining_seconds"],
        }
    )


@login_required
def chat_history(request):
    """
    GET endpoint to fetch recent chat messages for a specific session.
    Expects ?session_id=123
    """
    session_id = request.GET.get("session_id")

    ban_status = get_current_ban_status(request.user)

    if not session_id:
        return JsonResponse(
            {
                "messages": [],
                "is_banned": ban_status["is_banned"],
                "ban_remaining_seconds": ban_status["ban_remaining_seconds"],
            }
        )

    recent_msgs = ChatMessage.objects.filter(
        session_id=session_id, session__user=request.user
    ).order_by("-sent_at")[:50]

    messages_list = []
    for msg in reversed(list(recent_msgs)):
        role = "user" if msg.sender == request.user else "assistant"
        messages_list.append(
            {
                "role": role,
                "content": msg.message,
                "sent_at": msg.sent_at.strftime("%I:%M %p"),
            }
        )

    return JsonResponse(
        {
            "messages": messages_list,
            "is_banned": ban_status["is_banned"],
            "ban_remaining_seconds": ban_status["ban_remaining_seconds"],
        }
    )


@login_required
@require_POST
def chat_clear(request):
    """
    POST endpoint to clear a specific chat session.
    Expects JSON: { "session_id": 123 }
    """
    try:
        data = json.loads(request.body)
        session_id = data.get("session_id")
        if session_id:
            ChatSession.objects.filter(id=session_id, user=request.user).delete()
            return JsonResponse({"success": True})
        return JsonResponse({"error": "session_id required"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def select_design_type(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    # Get active packages to display as options
    active_packages = Package.objects.filter(is_active=True).order_by("price")

    return render(
        request, "client/select_design_type.html", {"packages": active_packages}
    )


@login_required
def my_designs_page(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    # Order by updated_at descending so newest are first
    designs_list = UserDesign.objects.filter(user=request.user).order_by("-updated_at")

    paginator = Paginator(designs_list, 8)  # Show 8 designs per page
    page_number = request.GET.get("page")
    designs = paginator.get_page(page_number)

    # Get active packages for the "Create New Design" modal
    active_packages = Package.objects.filter(is_active=True).order_by("price")

    return render(
        request,
        "client/my_designs.html",
        {"designs": designs, "packages": active_packages},
    )


@login_required
@require_POST
def save_user_design(request):
    if request.user.role != "customer":
        return JsonResponse({"status": "error", "message": "Not allowed"}, status=403)

    try:
        data = json.loads(request.body)
        design_id = data.get("id")
        name = data.get("name", "Untitled Design")
        canvas_json = data.get("canvas_json")
        thumbnail_data = data.get("thumbnail")  # Base64 string
        base_package_id = data.get("base_package_id")

        if not canvas_json:
            return JsonResponse(
                {"status": "error", "message": "Canvas data is required"}, status=400
            )

        # Handle thumbnail image (Base64)
        import base64
        import uuid

        from django.core.files.base import ContentFile

        image_file = None
        if thumbnail_data and "," in thumbnail_data:
            format, imgstr = thumbnail_data.split(";base64,")
            ext = format.split("/")[-1]
            image_file = ContentFile(
                base64.b64decode(imgstr), name=f"{uuid.uuid4().hex}.{ext}"
            )

        if design_id:
            # Update existing
            design = get_object_or_404(UserDesign, id=design_id, user=request.user)
            if name:
                design.name = name
            design.canvas_json = canvas_json
            if image_file:
                design.thumbnail = image_file
            
            # Maintain or update base package
            if base_package_id:
                try:
                    design.base_package = Package.objects.get(id=base_package_id)
                except Package.DoesNotExist:
                    pass
            
            design.save()
            log_action(request.user, f"Updated custom design #{design.id}.")
        else:
            # Create new
            base_package = None
            if base_package_id:
                try:
                    base_package = Package.objects.get(id=base_package_id)
                except Package.DoesNotExist:
                    pass

            design = UserDesign.objects.create(
                user=request.user,
                name=name,
                canvas_json=canvas_json,
                thumbnail=image_file,
                base_package=base_package,
            )
            log_action(request.user, f"Created new custom design #{design.id}.")

        return JsonResponse(
            {
                "status": "success",
                "id": design.id,
                "message": "Design saved successfully",
            }
        )
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
@require_POST
def rename_user_design(request, id):
    if request.user.role != "customer":
        return JsonResponse({"status": "error", "message": "Not allowed"}, status=403)

    try:
        design = get_object_or_404(UserDesign, id=id, user=request.user)
        data = json.loads(request.body)
        new_name = data.get("name")

        if not new_name or not new_name.strip():
            return JsonResponse(
                {"status": "error", "message": "Name cannot be empty"}, status=400
            )

        design.name = new_name.strip()
        design.save()
        log_action(
            request.user, f"Renamed custom design #{design.id} to '{design.name}'."
        )
        return JsonResponse({"status": "success", "name": design.name})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
@require_POST
def delete_user_design(request, id):
    if request.user.role != "customer":
        return JsonResponse({"status": "error", "message": "Not allowed"}, status=403)

    try:
        design = get_object_or_404(UserDesign, id=id, user=request.user)
        design_id_val = design.id
        design.delete()
        log_action(request.user, f"Deleted custom design #{design_id_val}.")
        return JsonResponse(
            {"status": "success", "message": "Design deleted successfully"}
        )
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
def design_canvas_page(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    context = {}

    # 1. Check if we're editing an existing design
    design_id = request.GET.get("id")
    package_id = request.GET.get("package_id")
    is_custom = request.GET.get("custom") == "true"

    if design_id:
        design = get_object_or_404(UserDesign, id=design_id, user=request.user)
        context["design"] = design
        if design.base_package:
            context["base_package"] = design.base_package
    elif package_id:
        # Starting a new design from a package
        base_package = get_object_or_404(Package, id=package_id)
        context["base_package"] = base_package
        # We don't save a UserDesign yet, we just pass the info to the frontend
    elif is_custom:
        # Starting a blank custom design
        pass
    else:
        # No ID and no package_id and no custom flag -> Redirect to selection
        return redirect("select_design_type")

    # 2. Extract quotas for all packages
    all_packages = Package.objects.all().order_by("price")
    all_package_quotas = {}
    all_categories = list(CanvasCategory.objects.values_list('name', flat=True))

    for pkg in all_packages:
        quotas = {}
        for feature in pkg.feature_list():
            feature_text = feature.strip()
            if not feature_text:
                continue

            # Case 1: "1 Backdrop", "50 Balloons"
            match = re.search(r"^(\d+)\s+(.+)$", feature_text, re.IGNORECASE)
            if match:
                qty = int(match.group(1))
                item_name = match.group(2).strip().lower()
                
                mapped_cat = None
                for cat_name in all_categories:
                    cat_lower = cat_name.lower()
                    if cat_lower in item_name or item_name in cat_lower or \
                       cat_lower.rstrip('s') in item_name or item_name.rstrip('s') in cat_lower:
                        mapped_cat = cat_lower
                        break
                
                if mapped_cat:
                    quotas[mapped_cat] = qty
                else:
                    quotas[item_name] = qty

                # NEW: Extract balloon color limit if present in this feature
                if "balloon" in feature_text.lower() and "color" in feature_text.lower():
                    color_match = re.search(r"max\s+(\d+)\s+colors", feature_text, re.IGNORECASE)
                    if color_match:
                        quotas["balloon_color_limit"] = int(color_match.group(1))
            else:
                feature_lower = feature_text.lower()
                
                # NEW: Extract balloon color limit for non-numeric features too
                if "balloon" in feature_lower and "color" in feature_lower:
                    color_match = re.search(r"max\s+(\d+)\s+colors", feature_text, re.IGNORECASE)
                    if color_match:
                        quotas["balloon_color_limit"] = int(color_match.group(1))

                mapped_cat = None
                for cat_name in all_categories:
                    cat_lower = cat_name.lower()
                    if cat_lower in feature_lower or feature_lower in cat_lower or \
                       cat_lower.rstrip('s') in feature_lower or feature_lower.rstrip('s') in cat_lower:
                        mapped_cat = cat_lower
                        break
                
                if mapped_cat:
                    quotas[mapped_cat] = 999
                else:
                    quotas[feature_text] = -1
        all_package_quotas[pkg.id] = quotas

    context["all_packages"] = all_packages
    context["all_package_quotas"] = json.dumps(all_package_quotas)

    # Set initial quotas for current base package
    if "base_package" in context:
        context["package_quotas"] = json.dumps(all_package_quotas.get(context["base_package"].id, {}))
    else:
        context["package_quotas"] = json.dumps({})

    # 3. Fetch AddOn prices to calculate visual cart
    addons = AddOn.objects.filter(is_active=True)
    addon_prices = {}
    for addon in addons:
        addon_prices[addon.name.lower()] = str(addon.price)

    context["addon_prices"] = json.dumps(addon_prices)

    categories = ["Backdrops", "Balloons", "Furniture", "Decorations"]
    context["categories"] = categories
    active_canvas_assets = list(
        CanvasAsset.objects.filter(is_active=True, category__is_active=True)
        .select_related("category")
        .order_by("category__order", "category__name", "label_ref__order", "subgroup", "sort_order", "id")
    )
    canvas_categories = list(CanvasCategory.objects.filter(is_active=True).order_by("order", "name"))

    assets_by_category = {}
    for asset in active_canvas_assets:
        assets_by_category.setdefault(asset.category_id, []).append(asset)

    # Attach explicit resolved assets per category so template rendering does not depend on prefetch state.
    for category in canvas_categories:
        category.design_assets = assets_by_category.get(category.id, [])

    context["canvas_categories"] = canvas_categories

    # Fallback payload for client-side hydration when sidebar cards are empty.
    canvas_assets_payload = {}
    for category in canvas_categories:
        category_key = category.name.lower().strip()
        items = []
        for asset in category.design_assets:
            src = ""
            if asset.image:
                src = asset.image.url
            elif asset.static_path:
                src = static(asset.static_path)
            items.append(
                {
                    "id": asset.id,
                    "label": asset.label or "Asset",
                    "src": src,
                    "category": category_key,
                    "type": asset.item_type or "image",
                    "width": int(asset.width or 150),
                    "height": int(asset.height or 150),
                }
            )
        canvas_assets_payload[category_key] = items
    context["canvas_assets_payload"] = json.dumps(canvas_assets_payload)
    context["canvas_assets_payload_obj"] = canvas_assets_payload

    return render(request, "client/design_canvas.html", context)


# =========================================
# ADMIN GALLERY MANAGEMENT
# =========================================


@login_required
def admin_gallery(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    categories = GalleryCategory.objects.all()
    selected_image_category = request.GET.get("image_category", "all")
    gallery_images = GalleryImage.objects.select_related("category").all()

    if selected_image_category != "all":
        try:
            category_id = int(selected_image_category)
            gallery_images = gallery_images.filter(category_id=category_id)
        except (TypeError, ValueError):
            selected_image_category = "all"

    paginator = Paginator(gallery_images, 5)
    page_number = request.GET.get("page")
    gallery_images = paginator.get_page(page_number)

    return render(
        request,
        "admin/gallery/admin_gallery.html",
        {
            "categories": categories,
            "gallery_images": gallery_images,
            "selected_image_category": selected_image_category,
        },
    )


@login_required
def admin_gallery_category_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        order = request.POST.get("order", 0)
        if not name:
            messages.error(request, "Category name is required.")
            return render(request, "admin/gallery/gallery_category_form.html")
        GalleryCategory.objects.create(name=name, order=int(order) if order else 0)
        log_action(request.user, f"Created gallery category '{name}'.")
        messages.success(request, "Category created successfully.")
        return redirect("admin_gallery")
    return render(request, "admin/gallery/gallery_category_form.html")


@login_required
def admin_gallery_category_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    category = get_object_or_404(GalleryCategory, id=id)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        order = request.POST.get("order", 0)
        if not name:
            messages.error(request, "Category name is required.")
            return render(
                request,
                "admin/gallery/gallery_category_form.html",
                {"category": category},
            )
        category.name = name
        category.order = int(order) if order else 0
        category.save()
        log_action(request.user, f"Updated gallery category '{name}'.")
        messages.success(request, "Category updated successfully.")
        return redirect("admin_gallery")
    return render(
        request, "admin/gallery/gallery_category_form.html", {"category": category}
    )


@login_required
def admin_gallery_category_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    category = get_object_or_404(GalleryCategory, id=id)
    cat_name = category.name
    category.delete()
    log_action(request.user, f"Deleted gallery category '{cat_name}'.")
    messages.success(request, "Category deleted successfully.")
    return redirect("admin_gallery")


@login_required
def admin_gallery_image_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    categories = GalleryCategory.objects.all()
    if request.method == "POST":
        category_id = request.POST.get("category")
        caption = request.POST.get("caption", "").strip()
        image = request.FILES.get("image")
        if not category_id or not image:
            messages.error(request, "Category and image are required.")
            return render(
                request,
                "admin/gallery/gallery_image_form.html",
                {"categories": categories},
            )
        category = get_object_or_404(GalleryCategory, id=category_id)
        GalleryImage.objects.create(category=category, image=image, caption=caption)
        log_action(request.user, f"Added gallery image to '{category.name}'.")
        messages.success(request, "Image added successfully.")
        return redirect("admin_gallery")
    return render(
        request, "admin/gallery/gallery_image_form.html", {"categories": categories}
    )


@login_required
def admin_gallery_image_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    gallery_image = get_object_or_404(GalleryImage, id=id)
    categories = GalleryCategory.objects.all()
    next_url = request.GET.get("next") or request.POST.get("next") or ""
    scroll_target = (
        request.GET.get("scroll_target") or request.POST.get("scroll_target") or ""
    )
    if request.method == "POST":
        category_id = request.POST.get("category")
        caption = request.POST.get("caption", "").strip()
        new_image = request.FILES.get("image")
        is_active = request.POST.get("is_active") == "on"
        if not category_id:
            messages.error(request, "Category is required.")
            return render(
                request,
                "admin/gallery/gallery_image_form.html",
                {
                    "gallery_image": gallery_image,
                    "categories": categories,
                    "next_url": next_url,
                    "scroll_target": scroll_target,
                },
            )
        gallery_image.category = get_object_or_404(GalleryCategory, id=category_id)
        gallery_image.caption = caption
        gallery_image.is_active = is_active
        if new_image:
            gallery_image.image = new_image
        gallery_image.save()
        log_action(request.user, f"Updated gallery image #{gallery_image.id}.")
        messages.success(request, "Image updated successfully.")
        if next_url.startswith("/"):
            if scroll_target:
                separator = "&" if "?" in next_url else "?"
                return redirect(f"{next_url}{separator}scroll={scroll_target}")
            return redirect(next_url)
        return redirect("admin_gallery")
    return render(
        request,
        "admin/gallery/gallery_image_form.html",
        {
            "gallery_image": gallery_image,
            "categories": categories,
            "next_url": next_url,
            "scroll_target": scroll_target,
        },
    )


@login_required
def admin_gallery_image_detail(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    gallery_image = get_object_or_404(
        GalleryImage.objects.select_related("category"), id=id
    )
    next_url = (
        request.GET.get("next") or "/staff/gallery/?scroll=gallery-images-section"
    )
    scroll_target = request.GET.get("scroll_target") or "gallery-images-section"
    if not str(next_url).startswith("/"):
        next_url = "/staff/gallery/?scroll=gallery-images-section"
    if scroll_target and "scroll=" not in next_url:
        separator = "&" if "?" in next_url else "?"
        next_url = f"{next_url}{separator}scroll={scroll_target}"

    return render(
        request,
        "admin/gallery/gallery_image_detail.html",
        {
            "gallery_image": gallery_image,
            "next_url": next_url,
        },
    )


@login_required
def admin_gallery_image_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    gallery_image = get_object_or_404(GalleryImage, id=id)
    img_id = gallery_image.id
    gallery_image.image.delete()  # Delete file from storage
    gallery_image.delete()
    log_action(request.user, f"Deleted gallery image #{img_id}.")
    messages.success(request, "Image deleted successfully.")
    return redirect("admin_gallery")


# =========================================
# ADMIN CANVAS ASSET MANAGEMENT
# =========================================


@login_required
def admin_canvas_assets(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    categories = CanvasCategory.objects.order_by("order", "name")
    selected_asset_category = request.GET.get("asset_category", "all")
    canvas_assets = CanvasAsset.objects.select_related("category").all()

    if selected_asset_category != "all":
        try:
            parsed_category_id = int(selected_asset_category)
            canvas_assets = canvas_assets.filter(category_id=parsed_category_id)
        except (TypeError, ValueError):
            selected_asset_category = "all"

    categories_paginator = Paginator(categories, 10)
    categories_page = categories_paginator.get_page(request.GET.get("cat_page"))

    assets_paginator = Paginator(canvas_assets, 8)
    canvas_assets = assets_paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "admin/canvas/admin_canvas_assets.html",
        {
            "categories": categories,
            "categories_page": categories_page,
            "canvas_assets": canvas_assets,
            "selected_asset_category": selected_asset_category,
        },
    )


@login_required
def admin_canvas_category_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        order = request.POST.get("order", 1)
        is_active = request.POST.get("is_active") == "on"
        if not name:
            messages.warning(request, "Please enter a category name before saving.")
            return render(request, "admin/canvas/canvas_category_form.html")
        CanvasCategory.objects.create(
            name=name, order=int(order) if order else 1, is_active=is_active
        )
        log_action(request.user, f"Created canvas category '{name}'.")
        messages.success(request, "Canvas category created successfully.")
        return redirect("admin_canvas_assets")
    return render(request, "admin/canvas/canvas_category_form.html")


@login_required
def admin_canvas_category_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    category = get_object_or_404(CanvasCategory, id=id)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        order = request.POST.get("order", 1)
        is_active = request.POST.get("is_active") == "on"
        if not name:
            messages.warning(request, "Please enter a category name before saving.")
            return render(
                request,
                "admin/canvas/canvas_category_form.html",
                {"category": category},
            )
        category.name = name
        category.order = int(order) if order else 1
        category.is_active = is_active
        category.save()
        log_action(request.user, f"Updated canvas category '{name}'.")
        messages.success(request, "Canvas category updated successfully.")
        return redirect("admin_canvas_assets")
    return render(
        request, "admin/canvas/canvas_category_form.html", {"category": category}
    )


@login_required
def admin_canvas_category_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    category = get_object_or_404(CanvasCategory, id=id)
    cat_name = category.name
    category.delete()
    log_action(request.user, f"Deleted canvas category '{cat_name}'.")
    messages.success(request, "Canvas category deleted successfully.")
    return redirect("admin_canvas_assets")


@login_required
def admin_canvas_label_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    categories = CanvasCategory.objects.order_by("order", "name")
    if request.method == "POST":
        category_id = request.POST.get("category")
        name = request.POST.get("name", "").strip()
        order = request.POST.get("order", 1)
        is_active = request.POST.get("is_active") == "on"
        if not category_id:
            messages.warning(request, "Please select a category for this label.")
            return render(
                request,
                "admin/canvas/canvas_label_form.html",
                {"categories": categories},
            )
        if not name:
            messages.warning(request, "Please enter a label name before saving.")
            return render(
                request,
                "admin/canvas/canvas_label_form.html",
                {"categories": categories},
            )

        category = CanvasCategory.objects.filter(id=category_id).first()
        if not category:
            messages.warning(
                request,
                "Selected category could not be found. Please choose another category.",
            )
            return render(
                request,
                "admin/canvas/canvas_label_form.html",
                {"categories": categories},
            )
        CanvasLabel.objects.create(
            category=category,
            name=name,
            order=int(order) if order else 1,
            is_active=is_active,
        )
        log_action(request.user, f"Created canvas label '{name}' in '{category.name}'.")
        messages.success(request, "Canvas label created successfully.")
        return redirect("admin_canvas_assets")
    return render(
        request, "admin/canvas/canvas_label_form.html", {"categories": categories}
    )


@login_required
def admin_canvas_label_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    label = get_object_or_404(CanvasLabel, id=id)
    categories = CanvasCategory.objects.order_by("order", "name")
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        order = request.POST.get("order", 1)
        is_active = request.POST.get("is_active") == "on"
        if not name:
            messages.warning(request, "Please enter a label name before saving.")
            return render(
                request,
                "admin/canvas/canvas_label_form.html",
                {"label": label, "categories": categories},
            )
        label.name = name
        label.order = int(order) if order else 1
        label.is_active = is_active
        label.save()
        log_action(request.user, f"Updated canvas label '{name}'.")
        messages.success(request, "Canvas label updated successfully.")
        return redirect("admin_canvas_assets")
    return render(
        request,
        "admin/canvas/canvas_label_form.html",
        {"label": label, "categories": categories},
    )


@login_required
def admin_canvas_label_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    label = get_object_or_404(CanvasLabel, id=id)
    label_name = label.name
    CanvasAsset.objects.filter(label_ref=label).update(label_ref=None)
    label.delete()
    log_action(request.user, f"Deleted canvas label '{label_name}'.")
    messages.success(request, "Canvas label deleted successfully.")
    return redirect("admin_canvas_assets")


@login_required
def admin_canvas_asset_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    categories = CanvasCategory.objects.order_by("order", "name")
    if request.method == "POST":
        category_id = request.POST.get("category")
        label = request.POST.get("label", "").strip()
        static_path = ""
        item_type = "image"
        image = request.FILES.get("image")
        width = request.POST.get("width", "150")
        height = request.POST.get("height", "150")
        is_active = request.POST.get("is_active") == "on"

        if not category_id:
            messages.warning(request, "Please select a category for this asset.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {"categories": categories},
            )

        if not label:
            messages.warning(request, "Please enter an asset name before saving.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {"categories": categories},
            )

        if not image:
            messages.warning(request, "Please upload an image file for this asset.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {"categories": categories},
            )

        try:
            width_val = int(width)
            height_val = int(height)
        except ValueError:
            messages.warning(
                request, "Default width and height must be valid whole numbers."
            )
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {"categories": categories},
            )

        if width_val < 1:
            messages.warning(request, "Default width must be at least 1.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {"categories": categories},
            )

        if height_val < 1:
            messages.warning(request, "Default height must be at least 1.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {"categories": categories},
            )

        category = CanvasCategory.objects.filter(id=category_id).first()
        if not category:
            messages.warning(
                request,
                "Selected category could not be found. Please choose another category.",
            )
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {"categories": categories},
            )

        next_sort = (
            CanvasAsset.objects.filter(category=category).aggregate(max_sort=Max("sort_order"))[
                "max_sort"
            ]
            or 0
        ) + 1
        CanvasAsset.objects.create(
            category=category,
            label_ref=None,
            label=label,
            subgroup="",
            static_path=static_path,
            item_type=item_type,
            image=image,
            width=width_val,
            height=height_val,
            sort_order=next_sort,
            is_active=is_active,
        )
        log_action(request.user, f"Added canvas asset '{label}' to '{category.name}'.")
        messages.success(request, "Canvas asset added successfully.")
        return redirect("admin_canvas_assets")

    return render(
        request,
        "admin/canvas/canvas_asset_form.html",
        {"categories": categories},
    )


@login_required
def admin_canvas_asset_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    canvas_asset = get_object_or_404(CanvasAsset, id=id)
    categories = CanvasCategory.objects.order_by("order", "name")
    next_url = request.GET.get("next") or request.POST.get("next") or ""
    scroll_target = (
        request.GET.get("scroll_target") or request.POST.get("scroll_target") or ""
    )

    if request.method == "POST":
        category_id = request.POST.get("category")
        label = request.POST.get("label", "").strip()
        static_path = canvas_asset.static_path or ""
        item_type = canvas_asset.item_type or "image"
        new_image = request.FILES.get("image")
        width = request.POST.get("width", "150")
        height = request.POST.get("height", "150")
        is_active = request.POST.get("is_active") == "on"

        if not category_id:
            messages.warning(request, "Please select a category for this asset.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {
                    "canvas_asset": canvas_asset,
                    "categories": categories,
                    "next_url": next_url,
                    "scroll_target": scroll_target,
                },
            )

        if not label:
            messages.warning(request, "Please enter an asset name before saving.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {
                    "canvas_asset": canvas_asset,
                    "categories": categories,
                    "next_url": next_url,
                    "scroll_target": scroll_target,
                },
            )

        if not new_image and not canvas_asset.image and not canvas_asset.static_path:
            messages.warning(request, "Please upload an image file for this asset.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {
                    "canvas_asset": canvas_asset,
                    "categories": categories,
                    "next_url": next_url,
                    "scroll_target": scroll_target,
                },
            )

        try:
            width_val = int(width)
            height_val = int(height)
        except ValueError:
            messages.warning(
                request, "Default width and height must be valid whole numbers."
            )
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {
                    "canvas_asset": canvas_asset,
                    "categories": categories,
                    "next_url": next_url,
                    "scroll_target": scroll_target,
                },
            )

        if width_val < 1:
            messages.warning(request, "Default width must be at least 1.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {
                    "canvas_asset": canvas_asset,
                    "categories": categories,
                    "next_url": next_url,
                    "scroll_target": scroll_target,
                },
            )

        if height_val < 1:
            messages.warning(request, "Default height must be at least 1.")
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {
                    "canvas_asset": canvas_asset,
                    "categories": categories,
                    "next_url": next_url,
                    "scroll_target": scroll_target,
                },
            )

        category = CanvasCategory.objects.filter(id=category_id).first()
        if not category:
            messages.warning(
                request,
                "Selected category could not be found. Please choose another category.",
            )
            return render(
                request,
                "admin/canvas/canvas_asset_form.html",
                {
                    "canvas_asset": canvas_asset,
                    "categories": categories,
                    "next_url": next_url,
                    "scroll_target": scroll_target,
                },
            )

        canvas_asset.category = category
        canvas_asset.label_ref = None
        canvas_asset.label = label
        canvas_asset.subgroup = ""
        canvas_asset.static_path = static_path
        canvas_asset.item_type = item_type
        canvas_asset.width = width_val
        canvas_asset.height = height_val
        canvas_asset.is_active = is_active
        if new_image:
            canvas_asset.image = new_image
        canvas_asset.save()

        log_action(request.user, f"Updated canvas asset #{canvas_asset.id}.")
        messages.success(request, "Canvas asset updated successfully.")
        if next_url.startswith("/"):
            if scroll_target:
                separator = "&" if "?" in next_url else "?"
                return redirect(f"{next_url}{separator}scroll={scroll_target}")
            return redirect(next_url)
        return redirect("admin_canvas_assets")

    return render(
        request,
        "admin/canvas/canvas_asset_form.html",
        {
            "canvas_asset": canvas_asset,
            "categories": categories,
            "next_url": next_url,
            "scroll_target": scroll_target,
        },
    )


@login_required
def admin_canvas_asset_detail(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    canvas_asset = get_object_or_404(
        CanvasAsset.objects.select_related("category", "label_ref"), id=id
    )
    next_url = (
        request.GET.get("next")
        or "/staff/canvas-assets/?scroll=canvas-assets-list-section"
    )
    scroll_target = request.GET.get("scroll_target") or "canvas-assets-list-section"
    if not str(next_url).startswith("/"):
        next_url = "/staff/canvas-assets/?scroll=canvas-assets-list-section"
    if scroll_target and "scroll=" not in next_url:
        separator = "&" if "?" in next_url else "?"
        next_url = f"{next_url}{separator}scroll={scroll_target}"

    return render(
        request,
        "admin/canvas/canvas_asset_detail.html",
        {
            "canvas_asset": canvas_asset,
            "next_url": next_url,
        },
    )


@login_required
def admin_canvas_asset_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    canvas_asset = get_object_or_404(CanvasAsset, id=id)
    asset_id = canvas_asset.id
    canvas_asset.delete()
    log_action(request.user, f"Deleted canvas asset #{asset_id}.")
    messages.success(request, "Canvas asset deleted successfully.")
    return redirect("admin_canvas_assets")


# =============================================================================
# PROFILE UPDATE
# =============================================================================


@login_required
def update_my_profile(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        phone_number = request.POST.get("phone_number", "").strip()

        user = request.user
        user.first_name = first_name
        user.last_name = last_name
        user.phone_number = phone_number
        user.save()
        messages.success(request, "Profile updated successfully.")
    return redirect("my_profile")


# =============================================================================
# MY PAYMENTS (Customer)
# =============================================================================


@login_required
def my_payments(request):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    user = request.user
    allowed_tabs = {"action_required", "remaining_balances", "payment_history"}
    active_tab = request.GET.get("tab", "action_required").strip()
    if active_tab not in allowed_tabs:
        active_tab = "action_required"

    # Payment history filters
    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()

    payments_qs = (
        Payment.objects.filter(booking__user=user)
        .select_related("booking")
        .prefetch_related("booking__images", "booking__payments")
        .order_by("-created_at")
    )

    if search_query:
        payments_qs = payments_qs.filter(
            Q(transaction_ref__icontains=search_query)
            | Q(gcash_reference_number__icontains=search_query)
            | Q(booking__id__icontains=search_query)
        )
    if status_filter:
        payments_qs = payments_qs.filter(payment_status=status_filter)

    pending_count = Payment.objects.filter(
        booking__user=user, payment_status="pending"
    ).count()
    verified_count = Payment.objects.filter(
        booking__user=user, payment_status="verified"
    ).count()
    rejected_count = Payment.objects.filter(
        booking__user=user, payment_status="rejected"
    ).count()
    total_records = payments_qs.count()
    total_pending_payment_count = pending_count

    payments_paginator = Paginator(payments_qs, 5)
    payments_page_number = request.GET.get("payments_page", 1)
    payments_page_obj = payments_paginator.get_page(payments_page_number)

    # Attach booking/payment summary fields for detailed tables and modal
    for pay in payments_page_obj:
        booking_payments = list(pay.booking.payments.all())
        verified_payments = [
            p for p in booking_payments if p.payment_status == "verified"
        ]
        verified_paid = sum(
            (p.amount for p in verified_payments),
            Decimal("0.00"),
        )
        booking_total = pay.booking.total_price or Decimal("0.00")
        booking_remaining = booking_total - verified_paid
        if booking_remaining < Decimal("0.00"):
            booking_remaining = Decimal("0.00")

        pay.booking.time_range_display = get_booking_time_range(pay.booking)
        pay.booking.total_verified_paid = verified_paid
        pay.booking.booking_remaining = booking_remaining
        pay.booking.total_payment_records = len(booking_payments)
        pay.booking.cleaned_special_requests = remove_end_time_tag(pay.booking.special_requests or "")

    # Action-required: approved bookings awaiting payment
    ar_search_query = request.GET.get("ar_search", "").strip()
    ar_action_filter = request.GET.get("ar_action", "").strip()

    all_action_bookings = (
        Booking.objects.filter(user=user, status__in=["pending_payment", "confirmed"])
        .prefetch_related("payments", "images")
        .order_by("-updated_at", "-id")
    )

    if ar_search_query:
        all_action_bookings = all_action_bookings.filter(
            Q(id__icontains=ar_search_query) | Q(event_type__icontains=ar_search_query)
        )

    # Filter bookings with remaining balance
    action_required_list = []
    partial_bookings = []  # Bookings with partial verified payments
    
    for b in all_action_bookings:
        booking_payments = list(b.payments.all())
        verified_payments = [p for p in booking_payments if p.payment_status == "verified"]
        verified_paid = sum((p.amount for p in verified_payments), Decimal("0.00"))
        remaining = (b.total_price or Decimal("0.00")) - verified_paid
        if remaining < Decimal("0.00"):
            remaining = Decimal("0.00")

        if remaining > Decimal("0.00"):
            b.time_range_display = get_booking_time_range(b)
            b.booking_remaining = remaining
            b.total_verified_paid = verified_paid
            b.total_payment_records = len(booking_payments)

            latest_payment = max(booking_payments, key=lambda p: p.created_at) if booking_payments else None
            latest_verified_payment = (
                max(
                    verified_payments,
                    key=lambda p: (p.paid_at or p.updated_at or p.created_at),
                )
                if verified_payments
                else None
            )

            if latest_verified_payment:
                b.latest_verified_paid_at = (
                    latest_verified_payment.paid_at
                    or latest_verified_payment.updated_at
                    or latest_verified_payment.created_at
                )
                b.latest_verified_paid_amount = latest_verified_payment.amount
            else:
                b.latest_verified_paid_at = None
                b.latest_verified_paid_amount = Decimal("0.00")

            b.last_payment_status_display = (
                latest_payment.get_payment_status_display()
                if latest_payment
                else "No Payment Yet"
            )
            b.last_payment_submitted_at = (
                latest_payment.created_at if latest_payment else None
            )

            # If customer already submitted a payment that is awaiting admin review,
            # this booking should no longer appear under "Action Required".
            if latest_payment and latest_payment.payment_status == "pending":
                continue
                
            if latest_payment and latest_payment.payment_status == "rejected":
                b.payment_action_display = "Re-upload"
                b.payment_action_disabled = False
                action_required_list.append(b)
            else:
                is_partial = verified_paid > Decimal("0.00")
                b.payment_action_display = "Pay Balance" if is_partial else "Pay Now"
                b.payment_action_disabled = False
                
                if is_partial:
                    partial_bookings.append(b)
                else:
                    action_required_list.append(b)
            
            b.cleaned_special_requests = remove_end_time_tag(b.special_requests or "")

    action_required_total_count = len(action_required_list)
    partial_total_count = len(partial_bookings)

    ar_paginator = Paginator(action_required_list, 5)
    ar_page_number = request.GET.get("ar_page", 1)
    action_required_page_obj = ar_paginator.get_page(ar_page_number)

    # Paginate partial bookings (Remaining Balances)
    partial_paginator = Paginator(partial_bookings, 5)
    partial_page_number = request.GET.get("partial_page", 1)
    partial_page_obj = partial_paginator.get_page(partial_page_number)

    return render(
        request,
        "client/my_payments.html",
        {
            "pending_count": pending_count,
            "verified_count": verified_count,
            "rejected_count": rejected_count,
            "total_pending_payment_count": total_pending_payment_count,
            "action_required_total_count": action_required_total_count,
            "partial_total_count": partial_total_count,
            "partial_page_obj": partial_page_obj,
            "total_records": total_records,
            "search_query": search_query,
            "status_filter": status_filter,
            "ar_search_query": ar_search_query,
            "ar_action_filter": ar_action_filter,
            "action_required_page_obj": action_required_page_obj,
            "payments_page_obj": payments_page_obj,
            "active_tab": active_tab,
        },
    )


# =============================================================================
# DOWNLOAD PAYMENT RECEIPT PDF (Customer)
# =============================================================================


@login_required
def download_payment_receipt_pdf(request, payment_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as rl_canvas

    payment = get_object_or_404(Payment, id=payment_id)

    # Customers may only download their own receipts; admins/staff may download any
    if request.user.role == "customer" and payment.booking.user != request.user:
        return HttpResponseForbidden("Not allowed")
    if request.user.role == "customer" and payment.payment_status != "verified":
        messages.warning(
            request,
            "Receipt is available only after your payment has been verified.",
        )
        return redirect("my_payments")

    buffer = io.BytesIO()
    p = rl_canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Title
    p.setFont("Helvetica-Bold", 18)
    p.drawString(50, height - 60, "Payment Receipt")
    p.setFont("Helvetica", 10)
    p.drawString(50, height - 80, "Balloorina.ph – Official Payment Receipt")

    # Divider
    p.setLineWidth(1)
    p.line(50, height - 90, width - 50, height - 90)

    y = height - 120
    line_height = 22

    def draw_row(label, value):
        nonlocal y
        p.setFont("Helvetica-Bold", 11)
        p.drawString(50, y, f"{label}:")
        p.setFont("Helvetica", 11)
        p.drawString(220, y, str(value))
        y -= line_height

    draw_row("Receipt #", f"PAY-{payment.id:06d}")
    draw_row("Booking ID", f"#{payment.booking.id}")
    draw_row(
        "Customer",
        payment.booking.user.get_full_name() or payment.booking.user.username,
    )
    draw_row("Amount", f"PHP {payment.amount:,.2f}")
    draw_row("Payment Method", payment.get_payment_method_display())
    draw_row("Payment Type", payment.get_payment_type_display())
    draw_row("Status", payment.get_payment_status_display())
    draw_row("Transaction Ref", payment.transaction_ref or "—")
    if payment.gcash_reference_number:
        draw_row("GCash Reference #", payment.gcash_reference_number)
    if payment.gcash_sender_name:
        draw_row("GCash Sender Name", payment.gcash_sender_name)
    draw_row("Date Submitted", payment.created_at.strftime("%B %d, %Y %I:%M %p"))
    if payment.paid_at:
        draw_row("Date Verified", payment.paid_at.strftime("%B %d, %Y %I:%M %p"))
    if payment.notes:
        draw_row("Admin Notes", payment.notes)

    y -= 20
    p.line(50, y + 10, width - 50, y + 10)
    p.setFont("Helvetica-Oblique", 9)
    p.drawString(
        50, y - 5, "This is an automatically generated receipt from Balloorina.ph."
    )
    p.drawString(
        50,
        y - 18,
        "For concerns, please contact us via our website or GCash-registered number.",
    )

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="receipt_PAY{payment.id:06d}.pdf"'
    )
    return response


# =============================================================================
# ADMIN SERVICE MANAGEMENT
# =============================================================================


@login_required
def admin_service_list(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    services = Service.objects.order_by("display_order")
    return render(
        request, "admin/service/admin_service_list.html", {"services": services}
    )


@login_required
def admin_service_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        description = request.POST.get("description", "").strip()
        features = request.POST.get("features", "").strip()
        is_active = request.POST.get("is_active") == "on"
        image = request.FILES.get("image")

        try:
            display_order = int(request.POST.get("display_order", 0))
        except (ValueError, TypeError):
            display_order = 0

        if not title or not description:
            messages.error(request, "Title and description are required.")
            return render(
                request,
                "admin/service/admin_service_form.html",
                {
                    "action": "Create",
                    "post_data": request.POST,
                },
            )

        service = Service(
            title=title,
            description=description,
            features=features,
            display_order=display_order,
            is_active=is_active,
        )
        if image:
            service.image = image
        service.save()

        log_action(
            request.user, f"Created service '{service.title}' (ID #{service.id})."
        )
        messages.success(request, f"Service '{service.title}' created successfully.")
        return redirect("admin_service_list")

    return render(
        request, "admin/service/admin_service_form.html", {"action": "Create"}
    )


@login_required
def admin_service_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    service = get_object_or_404(Service, id=id)

    if request.method == "POST":
        service.title = request.POST.get("title", service.title).strip()
        service.description = request.POST.get(
            "description", service.description
        ).strip()
        service.features = request.POST.get("features", service.features).strip()
        service.is_active = request.POST.get("is_active") == "on"

        try:
            service.display_order = int(
                request.POST.get("display_order", service.display_order)
            )
        except (ValueError, TypeError):
            pass

        if request.FILES.get("image"):
            service.image = request.FILES["image"]

        service.save()
        log_action(
            request.user, f"Updated service '{service.title}' (ID #{service.id})."
        )
        messages.success(request, f"Service '{service.title}' updated successfully.")
        return redirect("admin_service_list")

    return render(
        request,
        "admin/service/admin_service_form.html",
        {
            "action": "Edit",
            "service": service,
        },
    )


@login_required
def admin_service_detail(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    service = get_object_or_404(Service, id=id)
    return render(
        request, "admin/service/admin_service_detail.html", {"service": service}
    )


@login_required
def admin_service_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    service = get_object_or_404(Service, id=id)
    service_title = service.title
    service.delete()
    log_action(request.user, f"Deleted service '{service_title}' (ID #{id}).")
    messages.success(request, f"Service '{service_title}' deleted successfully.")
    return redirect("admin_service_list")


# =============================================================================
# ADMIN HOME CONTENT MANAGEMENT
# =============================================================================


@login_required
def admin_home_content(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    content = HomeContent.objects.first()
    if content is None:
        content = HomeContent.objects.create()

    if request.method == "POST":
        content.hero_title = request.POST.get("hero_title", content.hero_title).strip()
        content.hero_subheadline = request.POST.get(
            "hero_subheadline", content.hero_subheadline
        ).strip()
        content.stat_events_styled = request.POST.get(
            "stat_events_styled", content.stat_events_styled
        ).strip()
        content.stat_rating = request.POST.get(
            "stat_rating", content.stat_rating
        ).strip()
        content.stat_satisfaction = request.POST.get(
            "stat_satisfaction", content.stat_satisfaction
        ).strip()
        content.stat_response_time = request.POST.get(
            "stat_response_time", content.stat_response_time
        ).strip()
        content.why_choose_title = request.POST.get(
            "why_choose_title", content.why_choose_title
        ).strip()
        content.why_choose_subtitle = request.POST.get(
            "why_choose_subtitle", content.why_choose_subtitle
        ).strip()

        if request.FILES.get("hero_main_image"):
            content.hero_main_image = request.FILES["hero_main_image"]
        if request.FILES.get("hero_float_bottom_image"):
            content.hero_float_bottom_image = request.FILES["hero_float_bottom_image"]
        if request.FILES.get("hero_float_top_image"):
            content.hero_float_top_image = request.FILES["hero_float_top_image"]

        content.save()
        log_action(request.user, "Updated Home page content.")
        messages.success(request, "Home content updated successfully.")
        return redirect("admin_home_content")

    features = HomeFeatureItem.objects.all()
    return render(
        request,
        "admin/content/home_content.html",
        {
            "content": content,
            "features": features,
        },
    )


@login_required
def admin_home_feature_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    content = HomeContent.objects.first()
    if content is None:
        content = HomeContent.objects.create()

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        description = request.POST.get("description", "").strip()
        icon_class = request.POST.get("icon_class", "fas fa-star").strip()
        is_active = request.POST.get("is_active") == "on"

        try:
            display_order = int(request.POST.get("display_order", 0))
        except (ValueError, TypeError):
            display_order = 0

        if not title:
            if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
                return JsonResponse({"ok": False, "message": "Title is required."}, status=400)
            messages.error(request, "Title is required.")
            return render(
                request,
                "admin/content/home_feature_form.html",
                {
                    "action": "Create",
                    "feature": {},
                    "post_data": request.POST,
                },
            )

        HomeFeatureItem.objects.create(
            home_content=content,
            title=title,
            description=description,
            icon_class=icon_class,
            display_order=display_order,
            is_active=is_active,
        )
        log_action(request.user, f"Created home feature item '{title}'.")
        
        if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
            return JsonResponse({"ok": True, "message": "Feature item created successfully."})
        
        messages.success(request, "Feature item created successfully.")
        return redirect("admin_home_content")

    return render(request, "admin/content/home_feature_form.html", {"action": "Create", "feature": {}, "post_data": {}})


@login_required
def admin_home_feature_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    feature = get_object_or_404(HomeFeatureItem, id=id)

    if request.method == "POST":
        feature.title = request.POST.get("title", feature.title).strip()
        feature.description = request.POST.get(
            "description", feature.description
        ).strip()
        feature.icon_class = request.POST.get("icon_class", feature.icon_class).strip()
        feature.is_active = request.POST.get("is_active") == "on"

        try:
            feature.display_order = int(
                request.POST.get("display_order", feature.display_order)
            )
        except (ValueError, TypeError):
            pass

        feature.save()
        log_action(
            request.user,
            f"Updated home feature item '{feature.title}' (ID #{feature.id}).",
        )
        
        if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
            return JsonResponse({"ok": True, "message": "Feature item updated successfully."})
        
        messages.success(request, "Feature item updated successfully.")
        return redirect("admin_home_content")

    return render(
        request,
        "admin/content/home_feature_form.html",
        {
            "action": "Edit",
            "feature": feature,
            "post_data": {},
        },
    )


@login_required
def admin_home_feature_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    feature = get_object_or_404(HomeFeatureItem, id=id)
    feature_title = feature.title
    feature.delete()
    log_action(request.user, f"Deleted home feature item '{feature_title}' (ID #{id}).")
    messages.success(request, "Feature item deleted successfully.")
    return redirect("admin_home_content")


# =============================================================================
# ADMIN ABOUT CONTENT MANAGEMENT
# =============================================================================


@login_required
def admin_about_content(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    content = AboutContent.objects.first()
    if content is None:
        content = AboutContent.objects.create()

    if request.method == "POST":
        content.hero_title = request.POST.get("hero_title", content.hero_title).strip()
        content.hero_subtitle = request.POST.get(
            "hero_subtitle", content.hero_subtitle
        ).strip()
        content.story_label = request.POST.get(
            "story_label", content.story_label
        ).strip()
        content.story_title = request.POST.get(
            "story_title", content.story_title
        ).strip()
        content.story_paragraph_1 = request.POST.get(
            "story_paragraph_1", content.story_paragraph_1
        ).strip()
        content.story_paragraph_2 = request.POST.get(
            "story_paragraph_2", content.story_paragraph_2
        ).strip()
        content.stat_events_styled = request.POST.get(
            "stat_events_styled", content.stat_events_styled
        ).strip()
        content.stat_year_founded = request.POST.get(
            "stat_year_founded", content.stat_year_founded
        ).strip()
        content.stat_satisfaction = request.POST.get(
            "stat_satisfaction", content.stat_satisfaction
        ).strip()
        content.mission_label = request.POST.get(
            "mission_label", content.mission_label
        ).strip()
        content.mission_title = request.POST.get(
            "mission_title", content.mission_title
        ).strip()
        content.mission_paragraph_1 = request.POST.get(
            "mission_paragraph_1", content.mission_paragraph_1
        ).strip()
        content.mission_paragraph_2 = request.POST.get(
            "mission_paragraph_2", content.mission_paragraph_2
        ).strip()
        content.values_title = request.POST.get(
            "values_title", content.values_title
        ).strip()
        content.values_subtitle = request.POST.get(
            "values_subtitle", content.values_subtitle
        ).strip()

        if request.FILES.get("story_image"):
            content.story_image = request.FILES["story_image"]
        if request.FILES.get("mission_image"):
            content.mission_image = request.FILES["mission_image"]

        content.save()
        log_action(request.user, "Updated About page content.")
        messages.success(request, "About content updated successfully.")
        return redirect("admin_about_content")

    values = AboutValueItem.objects.all()
    return render(
        request,
        "admin/content/about_content.html",
        {
            "content": content,
            "values": values,
        },
    )


@login_required
def admin_about_value_create(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    content = AboutContent.objects.first()
    if content is None:
        content = AboutContent.objects.create()

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        description = request.POST.get("description", "").strip()
        icon_class = request.POST.get("icon_class", "fas fa-star").strip()
        is_active = request.POST.get("is_active") == "on"

        try:
            display_order = int(request.POST.get("display_order", 0))
        except (ValueError, TypeError):
            display_order = 0

        if not title:
            if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
                return JsonResponse({"ok": False, "message": "Title is required."}, status=400)
            messages.error(request, "Title is required.")
            return render(
                request,
                "admin/content/about_value_form.html",
                {
                    "action": "Create",
                    "post_data": request.POST,
                },
            )

        AboutValueItem.objects.create(
            about_content=content,
            title=title,
            description=description,
            icon_class=icon_class,
            display_order=display_order,
            is_active=is_active,
        )
        log_action(request.user, f"Created about value item '{title}'.")
        
        if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
            return JsonResponse({"ok": True, "message": "Value item created successfully."})
        
        messages.success(request, "Value item created successfully.")
        return redirect("admin_about_content")

    return render(request, "admin/content/about_value_form.html", {"action": "Create"})


@login_required
def admin_about_value_edit(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    value_item = get_object_or_404(AboutValueItem, id=id)

    if request.method == "POST":
        value_item.title = request.POST.get("title", value_item.title).strip()
        value_item.description = request.POST.get(
            "description", value_item.description
        ).strip()
        value_item.icon_class = request.POST.get(
            "icon_class", value_item.icon_class
        ).strip()
        value_item.is_active = request.POST.get("is_active") == "on"

        try:
            value_item.display_order = int(
                request.POST.get("display_order", value_item.display_order)
            )
        except (ValueError, TypeError):
            pass

        value_item.save()
        log_action(
            request.user,
            f"Updated about value item '{value_item.title}' (ID #{value_item.id}).",
        )
        
        if request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest":
            return JsonResponse({"ok": True, "message": "Value item updated successfully."})
        
        messages.success(request, "Value item updated successfully.")
        return redirect("admin_about_content")

    return render(
        request,
        "admin/content/about_value_form.html",
        {
            "action": "Edit",
            "value_item": value_item,
        },
    )


@login_required
def admin_about_value_delete(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    value_item = get_object_or_404(AboutValueItem, id=id)
    item_title = value_item.title
    value_item.delete()
    log_action(request.user, f"Deleted about value item '{item_title}' (ID #{id}).")
    messages.success(request, "Value item deleted successfully.")
    return redirect("admin_about_content")


# =============================================================================
# ADMIN REVIEWS
# =============================================================================


@login_required
def admin_reviews(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    search_query = request.GET.get("search", "").strip()
    rating_filter = request.GET.get("rating", "").strip()
    sort_filter = request.GET.get("sort", "-created_at").strip()

    reviews_qs = Review.objects.select_related("user", "booking").prefetch_related(
        "images"
    )

    if search_query:
        reviews_qs = reviews_qs.filter(
            Q(user__first_name__icontains=search_query)
            | Q(user__last_name__icontains=search_query)
            | Q(user__username__icontains=search_query)
            | Q(comment__icontains=search_query)
        )

    if rating_filter:
        try:
            reviews_qs = reviews_qs.filter(rating=int(rating_filter))
        except (ValueError, TypeError):
            pass

    valid_sorts = ["-created_at", "created_at", "-rating", "rating"]
    if sort_filter not in valid_sorts:
        sort_filter = "-created_at"
    reviews_qs = reviews_qs.order_by(sort_filter)

    total_reviews = Review.objects.count()
    avg_rating_data = Review.objects.aggregate(avg=Avg("rating"))
    avg_rating = round(avg_rating_data["avg"] or 0, 1)
    featured_count = Review.objects.filter(is_testimonial=True).count()

    paginator = Paginator(reviews_qs, 15)
    page_number = request.GET.get("page", 1)
    reviews_page = paginator.get_page(page_number)

    return render(
        request,
        "admin/admin_reviews.html",
        {
            "total_reviews": total_reviews,
            "avg_rating": avg_rating,
            "featured_count": featured_count,
            "search_query": search_query,
            "rating_filter": rating_filter,
            "sort_filter": sort_filter,
            "reviews": reviews_page,
        },
    )


@login_required
def admin_review_toggle_testimonial(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    review = get_object_or_404(Review, id=id)
    review.is_testimonial = not review.is_testimonial
    review.save()

    status_label = (
        "featured as testimonial"
        if review.is_testimonial
        else "removed from testimonials"
    )
    log_action(request.user, f"Review #{review.id} {status_label}.")
    messages.success(request, f"Review #{review.id} {status_label}.")
    return redirect("admin_reviews")


# =============================================================================
# ADMIN CONCERNS
# =============================================================================


@login_required
def admin_concerns(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    context = build_concerns_context(request)
    return render(request, "admin/admin_concerns.html", context)


# =============================================================================
# ADMIN ANALYTICS
# =============================================================================


@login_required
def admin_analytics(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")
    return render(request, "admin/admin_analytics.html", build_dashboard_context(request))


@login_required
def admin_analytics_export_excel(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    from openpyxl.styles import Alignment, Font, PatternFill

    start_date, end_date = _get_reporting_date_range(request)
    selected_event_type = (request.GET.get("event_type") or "all").strip()

    bookings = Booking.objects.select_related("user").filter(
        created_at__date__gte=start_date, created_at__date__lte=end_date
    )
    if selected_event_type != "all":
        bookings = bookings.filter(event_type=selected_event_type)
    bookings = bookings.order_by("-created_at")

    payments = Payment.objects.select_related("booking").filter(
        created_at__date__gte=start_date, created_at__date__lte=end_date
    )
    if selected_event_type != "all":
        payments = payments.filter(booking__event_type=selected_event_type)
    payments = payments.order_by("-created_at")

    reviews = Review.objects.select_related("user").filter(
        created_at__date__gte=start_date, created_at__date__lte=end_date
    )
    if selected_event_type != "all":
        reviews = reviews.filter(booking__event_type=selected_event_type)
    reviews = reviews.order_by("-created_at")

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F46E5", end_color="4F46E5", fill_type="solid"
    )
    center_align = Alignment(horizontal="center")

    # ── Sheet 1: Bookings Summary ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Bookings Summary"

    booking_headers = [
        "Booking ID",
        "Customer",
        "Event Type",
        "Event Date",
        "Status",
        "Total Price (PHP)",
        "Payment Status",
    ]
    for col_idx, h in enumerate(booking_headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_idx, booking in enumerate(bookings, 2):
        ws1.cell(row=row_idx, column=1, value=booking.id)
        ws1.cell(
            row=row_idx,
            column=2,
            value=booking.user.get_full_name() or booking.user.username,
        )
        ws1.cell(row=row_idx, column=3, value=booking.event_type or "—")
        ws1.cell(
            row=row_idx,
            column=4,
            value=str(booking.event_date) if booking.event_date else "—",
        )
        ws1.cell(row=row_idx, column=5, value=booking.status)
        ws1.cell(row=row_idx, column=6, value=float(booking.total_price or 0))
        ws1.cell(row=row_idx, column=7, value=booking.payment_status or "—")

    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Sheet 2: Revenue Data ──────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Revenue Data")

    rev_headers = [
        "Payment ID",
        "Booking ID",
        "Amount (PHP)",
        "Method",
        "Type",
        "Status",
        "Date Submitted",
    ]
    for col_idx, h in enumerate(rev_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_idx, pay in enumerate(payments, 2):
        ws2.cell(row=row_idx, column=1, value=pay.id)
        ws2.cell(row=row_idx, column=2, value=pay.booking.id)
        ws2.cell(row=row_idx, column=3, value=float(pay.amount or 0))
        ws2.cell(row=row_idx, column=4, value=pay.get_payment_method_display())
        ws2.cell(row=row_idx, column=5, value=pay.get_payment_type_display())
        ws2.cell(row=row_idx, column=6, value=pay.get_payment_status_display())
        ws2.cell(
            row=row_idx,
            column=7,
            value=pay.created_at.strftime("%Y-%m-%d %H:%M") if pay.created_at else "—",
        )

    for col in ws2.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Sheet 3: Reviews Summary ───────────────────────────────────────────
    ws3 = wb.create_sheet(title="Reviews")

    review_headers = [
        "Review ID",
        "Customer",
        "Rating",
        "Comment",
        "Is Testimonial",
        "Date",
    ]
    for col_idx, h in enumerate(review_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_idx, rev in enumerate(reviews, 2):
        ws3.cell(row=row_idx, column=1, value=rev.id)
        ws3.cell(
            row=row_idx, column=2, value=rev.user.get_full_name() or rev.user.username
        )
        ws3.cell(row=row_idx, column=3, value=rev.rating)
        ws3.cell(row=row_idx, column=4, value=rev.comment[:200] if rev.comment else "")
        ws3.cell(row=row_idx, column=5, value="Yes" if rev.is_testimonial else "No")
        ws3.cell(
            row=row_idx,
            column=6,
            value=rev.created_at.strftime("%Y-%m-%d")
            if hasattr(rev, "created_at") and rev.created_at
            else "—",
        )

    for col in ws3.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws3.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"analytics_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    log_action(request.user, "Exported analytics data to Excel.")
    return response


@login_required
def admin_analytics_export_pdf(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    context = build_dashboard_context(request)
    
    # Ensure start_date and end_date are date objects for the template's |date filter
    from datetime import datetime
    try:
        context['start_date'] = datetime.strptime(context['start_date'], "%Y-%m-%d")
        context['end_date'] = datetime.strptime(context['end_date'], "%Y-%m-%d")
    except (ValueError, TypeError):
        pass
        
    context['timezone'] = timezone
    from django.conf import settings
    context['STATIC_ROOT'] = settings.STATIC_ROOT or settings.BASE_DIR / 'static'

    template = get_template("admin/analytics_pdf_template.html")
    html = template.render(context, request)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="analytics_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse(
            "Error generating analytics PDF. Please try again.", status=500
        )

    log_action(request.user, "Exported analytics data to PDF.")
    return response


# =============================================================================
# PAYMENT PAGE (Customer – GCash upload flow)
# =============================================================================


def _get_booking_payment_breakdown(booking):
    config = GCashConfig.objects.first()
    dp_percent = config.downpayment_percent if config else 20
    total_price = booking.total_price or Decimal("0.00")
    verified_paid = booking.payments.filter(payment_status="verified").aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")
    remaining_balance = total_price - verified_paid
    dp_amount = (total_price * Decimal(dp_percent) / Decimal(100)).quantize(
        Decimal("0.01")
    )
    return config, Decimal(dp_percent), total_price, verified_paid, remaining_balance, dp_amount


def _refresh_booking_payment_status(booking):
    total_paid = booking.payments.filter(payment_status="verified").aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")
    total_price = booking.total_price or Decimal("0.00")
    if total_paid >= total_price:
        booking.payment_status = "paid"
    elif total_paid > Decimal("0.00"):
        booking.payment_status = "partial"
    else:
        booking.payment_status = "pending"
    booking.save(update_fields=["payment_status"])
    return total_paid, total_price - total_paid


def _repair_legacy_auto_verified_paymongo():
    """
    Safety net:
    Legacy PayMongo flow could mark payments as verified before admin action.
    Any PayMongo payment with verified status but no verifying admin is treated as pending.
    """
    suspicious_ids = list(
        Payment.objects.filter(
            payment_method__startswith="paymongo_",
            payment_status="verified",
            verified_by__isnull=True,
        ).values_list("id", flat=True)
    )

    if not suspicious_ids:
        return

    affected_booking_ids = list(
        Payment.objects.filter(id__in=suspicious_ids)
        .values_list("booking_id", flat=True)
        .distinct()
    )

    Payment.objects.filter(id__in=suspicious_ids).update(
        payment_status="pending",
        paid_at=None,
    )

    for booking_id in affected_booking_ids:
        booking = Booking.objects.filter(id=booking_id).first()
        if booking:
            _refresh_booking_payment_status(booking)


@login_required
def payment_page(request, booking_id):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    _repair_legacy_auto_verified_paymongo()

    booking = get_object_or_404(Booking, id=booking_id, user=request.user)

    if booking.status not in ["pending_payment", "confirmed"]:
        messages.warning(
            request, "This booking is not currently available for payment."
        )
        return redirect("customer_profile")

    payment_history = booking.payments.select_related("verified_by").order_by(
        "-created_at"
    )
    (
        config,
        dp_percent,
        total_price,
        total_paid,
        remaining_balance,
        dp_amount,
    ) = _get_booking_payment_breakdown(booking)

    has_downpayment = payment_history.filter(
        payment_status="verified", payment_type="downpayment"
    ).exists()
    has_full_payment = payment_history.filter(
        payment_status="verified", payment_type="full"
    ).exists()
    is_fully_paid = remaining_balance <= Decimal("0.00")

    pending_payment = payment_history.filter(payment_status="pending").first()
    rejected_payment = (
        payment_history.filter(payment_status="rejected")
        .order_by("-updated_at")
        .first()
    )
    paymongo_method_types = ["gcash"] if settings.PAYMONGO_SECRET_KEY else []
    is_initial_payment = total_paid <= Decimal("0.00")
    downpayment_due = dp_amount if is_initial_payment else Decimal("0.00")
    full_amount_due = remaining_balance

    return render(
        request,
        "client/payment_upload.html",
        {
            "booking": booking,
            "is_fully_paid": is_fully_paid,
            "pending_payment": pending_payment,
            "pending_checkout_url": None,
            "rejected_payment": rejected_payment,
            "failed_payment": None,
            "has_downpayment": has_downpayment,
            "has_full_payment": has_full_payment,
            "dp_percent": dp_percent,
            "dp_amount": dp_amount,
            "total_price": total_price,
            "total_paid": total_paid,
            "remaining_balance": remaining_balance,
            "paymongo_method_types": paymongo_method_types,
            "payment_history": payment_history,
            "config": config,
            "is_initial_payment": is_initial_payment,
            "downpayment_due": downpayment_due,
            "full_amount_due": full_amount_due,
        },
    )


@login_required
def submit_payment(request, booking_id):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    booking = get_object_or_404(Booking, id=booking_id, user=request.user)

    if booking.status not in ["pending_payment", "confirmed"]:
        messages.error(request, "This booking is not available for payment.")
        return redirect("customer_profile")

    if request.method != "POST":
        return redirect("payment_page", booking_id=booking_id)

    gcash_ref_number = request.POST.get("gcash_ref_number", "").strip()
    sender_name = request.POST.get("sender_name", "").strip()
    amount_str = request.POST.get("amount", "").strip()
    payment_option = request.POST.get("payment_option", "").strip().lower()
    refund_ack = request.POST.get("refund_ack")
    receipt_image = request.FILES.get("receipt_image")

    if not gcash_ref_number or not sender_name or not amount_str:
        messages.error(request, "Please fill in all required payment fields.")
        return redirect("payment_page", booking_id=booking_id)
    if not refund_ack:
        messages.error(
            request,
            "Please confirm the non-refundable payment reminder before submitting.",
        )
        return redirect("payment_page", booking_id=booking_id)

    try:
        amount = Decimal(amount_str)
        if amount <= 0:
            raise ValueError("Amount must be positive.")
    except (InvalidOperation, ValueError):
        messages.error(request, "Invalid payment amount. Please enter a valid number.")
        return redirect("payment_page", booking_id=booking_id)

    # Block duplicate pending submissions
    if booking.payments.filter(payment_status="pending").exists():
        messages.warning(
            request,
            "You already have a payment pending review. Please wait for it to be processed.",
        )
        return redirect("payment_page", booking_id=booking_id)

    (
        _config,
        _dp_percent,
        total_price,
        verified_paid,
        remaining_balance,
        dp_amount,
    ) = _get_booking_payment_breakdown(booking)

    if amount > remaining_balance:
        messages.error(
            request,
            f"Amount exceeds remaining balance of PHP {remaining_balance:,.2f}.",
        )
        return redirect("payment_page", booking_id=booking_id)

    expected_amount = Decimal("0.00")
    if verified_paid > Decimal("0.00"):
        payment_type = "balance"
        expected_amount = remaining_balance.quantize(Decimal("0.01"))
    else:
        if payment_option not in {"downpayment", "full"}:
            messages.error(request, "Please select Downpayment or Full Payment.")
            return redirect("payment_page", booking_id=booking_id)
        if payment_option == "downpayment":
            payment_type = "downpayment"
            expected_amount = dp_amount.quantize(Decimal("0.01"))
        else:
            payment_type = "full"
            expected_amount = total_price.quantize(Decimal("0.01"))

    if amount.quantize(Decimal("0.01")) != expected_amount:
        messages.error(
            request,
            f"Invalid amount for selected option. Expected PHP {expected_amount:,.2f}.",
        )
        return redirect("payment_page", booking_id=booking_id)

    if payment_type == "downpayment" and amount >= total_price:
        messages.error(
            request,
            "Downpayment amount cannot be the same as full payment amount.",
        )
        return redirect("payment_page", booking_id=booking_id)

    transaction_ref = str(uuid.uuid4()).replace("-", "")[:16].upper()

    payment = Payment(
        booking=booking,
        amount=amount,
        payment_method="gcash",
        payment_type=payment_type,
        payment_status="pending",
        transaction_ref=transaction_ref,
        gcash_reference_number=gcash_ref_number,
        gcash_sender_name=sender_name,
    )
    if receipt_image:
        payment.receipt_image = receipt_image
    payment.save()

    # Notify admin/staff of the new payment
    AdminNotification.objects.create(
        booking=booking,
        user=request.user,
        message=(
            f"New GCash payment submitted for Booking #{booking.id} "
            f"by {request.user.get_full_name() or request.user.username}. "
            f"Amount: PHP {amount:,.2f} | Ref: {gcash_ref_number}"
        ),
    )

    log_action(
        request.user,
        f"Submitted GCash payment (txn: {transaction_ref}, ref: {gcash_ref_number}) for Booking #{booking.id}.",
    )
    messages.success(
        request, "Payment submitted successfully! We will verify it within 24 hours."
    )
    return redirect("payment_success", booking_id=booking_id)


@login_required
def payment_success(request, booking_id):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    _repair_legacy_auto_verified_paymongo()

    booking = get_object_or_404(Booking, id=booking_id, user=request.user)

    payment_history = booking.payments.select_related("verified_by").order_by(
        "-created_at"
    )
    (
        config,
        dp_percent,
        total_price,
        total_paid,
        remaining_balance,
        dp_amount,
    ) = _get_booking_payment_breakdown(booking)

    has_downpayment = payment_history.filter(
        payment_status="verified", payment_type="downpayment"
    ).exists()
    has_full_payment = payment_history.filter(
        payment_status="verified", payment_type="full"
    ).exists()
    is_fully_paid = remaining_balance <= Decimal("0.00")

    pending_payment = payment_history.filter(payment_status="pending").first()
    rejected_payment = (
        payment_history.filter(payment_status="rejected")
        .order_by("-updated_at")
        .first()
    )
    paymongo_method_types = ["gcash"] if settings.PAYMONGO_SECRET_KEY else []
    is_initial_payment = total_paid <= Decimal("0.00")
    downpayment_due = dp_amount if is_initial_payment else Decimal("0.00")
    full_amount_due = remaining_balance

    return render(
        request,
        "client/payment_upload.html",
        {
            "booking": booking,
            "is_fully_paid": is_fully_paid,
            "pending_payment": pending_payment,
            "pending_checkout_url": None,
            "rejected_payment": rejected_payment,
            "failed_payment": None,
            "has_downpayment": has_downpayment,
            "has_full_payment": has_full_payment,
            "dp_percent": dp_percent,
            "dp_amount": dp_amount,
            "total_price": total_price,
            "total_paid": total_paid,
            "remaining_balance": remaining_balance,
            "paymongo_method_types": paymongo_method_types,
            "payment_history": payment_history,
            "config": config,
            "payment_submitted": True,
            "is_initial_payment": is_initial_payment,
            "downpayment_due": downpayment_due,
            "full_amount_due": full_amount_due,
        },
    )


@login_required
def payment_cancel(request, booking_id):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    messages.info(request, "Payment was cancelled. You can try again at any time.")
    return redirect("payment_page", booking_id=booking.id)


# =============================================================================
# ADMIN PAYMENT MANAGEMENT
# =============================================================================


@login_required
def admin_payment_list(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    _repair_legacy_auto_verified_paymongo()

    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()
    type_filter = request.GET.get("type", "").strip()

    payments_qs = Payment.objects.select_related(
        "booking", "booking__user", "verified_by"
    ).order_by("-created_at")

    if search_query:
        payments_qs = payments_qs.filter(
            Q(transaction_ref__icontains=search_query)
            | Q(gcash_reference_number__icontains=search_query)
            | Q(gcash_sender_name__icontains=search_query)
            | Q(booking__user__first_name__icontains=search_query)
            | Q(booking__user__last_name__icontains=search_query)
            | Q(booking__user__username__icontains=search_query)
        )

    if status_filter:
        payments_qs = payments_qs.filter(payment_status=status_filter)

    if type_filter:
        payments_qs = payments_qs.filter(payment_type=type_filter)

    pending_count = Payment.objects.filter(payment_status="pending").count()
    verified_count = Payment.objects.filter(payment_status="verified").count()
    rejected_count = Payment.objects.filter(payment_status="rejected").count()

    total_revenue = Payment.objects.filter(payment_status="verified").aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    total_payment_records = payments_qs.count()

    # Build list of bookings where downpayment is not yet fully paid
    # Attach template-expected attributes: booking_id, customer_name, username,
    # total_paid, remaining_balance so the template can use them directly.
    gcash_config = GCashConfig.objects.first()
    dp_percent = gcash_config.downpayment_percent if gcash_config else 20
    
    all_active_bookings = (
        Booking.objects.filter(status__in=["pending_payment", "confirmed"])
        .select_related("user")
        .prefetch_related("payments")
        .order_by("-id")
    )

    balance_bookings = []
    for b in all_active_bookings:
        verified_paid = b.payments.filter(payment_status="verified").aggregate(
            total=Sum("amount")
        )["total"] or Decimal("0.00")
        
        total_price = b.total_price or Decimal("0.00")
        required_downpayment = (total_price * Decimal(dp_percent) / Decimal(100)).quantize(Decimal("0.01"))
        
        if verified_paid < required_downpayment:
            remaining_dp = required_downpayment - verified_paid
            b.booking_id = b.id
            b.customer_name = b.user.get_full_name() or b.user.username
            b.username = b.user.username
            b.total_paid = verified_paid
            b.remaining_balance = remaining_dp
            b.outstanding_balance = remaining_dp
            balance_bookings.append(b)

    bookings_with_balance_count = len(balance_bookings)

    # Pagination for payments table (8 items per page)
    paginator = Paginator(payments_qs, 8)
    page_number = request.GET.get("page", 1)
    payments_page = paginator.get_page(page_number)
    
    # Pagination for downpayment balance table (8 items per page)
    balance_paginator = Paginator(balance_bookings, 8)
    balance_page_number = request.GET.get("balance_page", 1)
    balance_page = balance_paginator.get_page(balance_page_number)

    return render(
        request,
        "admin/payment/admin_payment_list.html",
        {
            "pending_count": pending_count,
            "verified_count": verified_count,
            "rejected_count": rejected_count,
            "total_revenue": total_revenue,
            "bookings_with_balance_count": bookings_with_balance_count,
            "total_payment_records": total_payment_records,
            "search_query": search_query,
            "status_filter": status_filter,
            "type_filter": type_filter,
            "payments": payments_page,
            "balance_bookings": balance_page,
        },
    )


@login_required
def admin_payment_detail(request, id):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    _repair_legacy_auto_verified_paymongo()

    payment = get_object_or_404(Payment, id=id)
    booking = payment.booking

    payment_history = booking.payments.select_related("verified_by").order_by(
        "-created_at"
    )
    total_paid = payment_history.filter(payment_status="verified").aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")
    remaining_balance = (booking.total_price or Decimal("0.00")) - total_paid

    return render(
        request,
        "admin/payment/admin_payment_detail.html",
        {
            "payment": payment,
            "booking": booking,
            "total_paid": total_paid,
            "remaining_balance": remaining_balance,
            "payment_history": payment_history,
        },
    )


@login_required
def admin_payment_action(request, id, action):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    if request.method != "POST":
        return redirect("admin_payment_detail", id=id)

    payment = get_object_or_404(Payment, id=id)
    booking = payment.booking

    if action == "verify":
        payment.payment_status = "verified"
        payment.paid_at = timezone.now()
        payment.verified_by = request.user
        payment.save()

        # Recalculate booking payment status.
        # Business rule: once admin verifies any customer payment,
        # booking should be marked as confirmed immediately.
        total_paid = booking.payments.filter(payment_status="verified").aggregate(
            total=Sum("amount")
        )["total"] or Decimal("0.00")
        total_price = booking.total_price or Decimal("0.00")

        if total_paid >= total_price:
            booking.payment_status = "paid"
        else:
            booking.payment_status = "partial"
        booking.status = "confirmed"
        booking.save()

        # Notify the customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=(
                f"Your payment of PHP {payment.amount:,.2f} for Booking #{booking.id} "
                f"has been verified. Thank you!"
            ),
        )

        log_action(
            request.user, f"Verified payment #{payment.id} for Booking #{booking.id}."
        )
        messages.success(
            request, f"Payment #{payment.id} has been verified successfully."
        )

    elif action == "reject":
        admin_notes = request.POST.get("admin_notes", "").strip()
        payment.payment_status = "rejected"
        payment.notes = admin_notes
        payment.save(update_fields=["payment_status", "notes", "updated_at"])

        # Keep booking in pending payment state when payment is rejected,
        # unless there are other verified payments already recorded.
        verified_total = booking.payments.filter(payment_status="verified").aggregate(
            total=Sum("amount")
        )["total"] or Decimal("0.00")
        total_price = booking.total_price or Decimal("0.00")
        if verified_total <= Decimal("0.00"):
            booking.payment_status = "pending"
            booking.status = "pending_payment"
        elif verified_total >= total_price:
            booking.payment_status = "paid"
            booking.status = "confirmed"
        else:
            booking.payment_status = "partial"
            booking.status = "confirmed"
        booking.save(update_fields=["payment_status", "status", "updated_at"])

        # Notify the customer
        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=(
                f"Your payment of PHP {payment.amount:,.2f} for Booking #{booking.id} "
                f"was rejected. Reason: {admin_notes or 'Please contact us for details.'}"
            ),
        )

        log_action(
            request.user,
            f"Rejected payment #{payment.id} for Booking #{booking.id}. Reason: {admin_notes}",
        )
        messages.warning(request, f"Payment #{payment.id} has been rejected.")

    else:
        messages.error(
            request, f"Unknown action: '{action}'. Expected 'verify' or 'reject'."
        )

    return redirect("admin_payment_detail", id=id)


@login_required
def admin_gcash_config(request):
    if request.user.role not in ["admin", "staff"]:
        return HttpResponseForbidden("Not allowed")

    config = GCashConfig.objects.first()
    if config is None:
        config = GCashConfig.objects.create()

    if request.method == "POST":
        config.gcash_number = request.POST.get(
            "gcash_number", config.gcash_number
        ).strip()
        config.gcash_name = request.POST.get("gcash_name", config.gcash_name).strip()
        config.instructions = request.POST.get(
            "instructions", config.instructions
        ).strip()

        try:
            dp_percent = int(
                request.POST.get("downpayment_percent", config.downpayment_percent)
            )
            if 1 <= dp_percent <= 100:
                config.downpayment_percent = dp_percent
            else:
                messages.warning(
                    request, "Downpayment percent must be between 1 and 100."
                )
        except (ValueError, TypeError):
            messages.warning(
                request,
                "Invalid downpayment percent value. Keeping the previous value.",
            )

        if request.FILES.get("qr_code_image"):
            config.qr_code_image = request.FILES["qr_code_image"]

        config.save()
        log_action(request.user, "Updated GCash configuration.")
        messages.success(request, "GCash configuration updated successfully.")
        return redirect("admin_gcash_config")

    return render(request, "admin/payment/admin_gcash_config.html", {"config": config})


# =============================================================================
# PAYMONGO INTEGRATION VIEWS
# =============================================================================


@login_required
def create_paymongo_checkout(request, booking_id):
    if request.user.role != "customer":
        return HttpResponseForbidden("Not allowed")

    booking = get_object_or_404(Booking, id=booking_id, user=request.user)

    if booking.status not in ["pending_payment", "confirmed"]:
        messages.warning(request, "This booking is not available for payment.")
        return redirect("customer_profile")

    if request.method != "POST":
        return redirect("payment_page", booking_id=booking.id)

    if not settings.PAYMONGO_SECRET_KEY:
        messages.error(
            request,
            "PayMongo is not configured yet. Please contact support.",
        )
        return redirect("payment_page", booking_id=booking.id)

    payment_method_type = request.POST.get("payment_method", "gcash").strip().lower()
    amount_str = request.POST.get("amount", "").strip()
    payment_option = request.POST.get("payment_option", "").strip().lower()
    refund_ack = request.POST.get("refund_ack")
    allowed_method_types = {"gcash"}

    if payment_method_type not in allowed_method_types:
        messages.error(request, "Invalid PayMongo payment method.")
        return redirect("payment_page", booking_id=booking.id)

    if not amount_str:
        messages.error(request, "Please enter an amount.")
        return redirect("payment_page", booking_id=booking.id)

    try:
        amount = Decimal(amount_str)
        if amount <= 0:
            raise ValueError()
    except (ValueError, InvalidOperation):
        messages.error(request, "Invalid amount.")
        return redirect("payment_page", booking_id=booking.id)

    if not refund_ack:
        messages.error(
            request,
            "Please confirm the non-refundable payment reminder before continuing.",
        )
        return redirect("payment_page", booking_id=booking.id)

    if booking.payments.filter(payment_status="pending").exists():
        messages.warning(
            request,
            "You already have a payment pending review. Please wait for it to be processed.",
        )
        return redirect("payment_page", booking_id=booking.id)

    (
        _config,
        _dp_percent,
        total_price,
        verified_paid,
        remaining_balance,
        dp_amount,
    ) = _get_booking_payment_breakdown(booking)

    if amount > remaining_balance:
        messages.error(
            request,
            f"Amount exceeds remaining balance of PHP {remaining_balance:,.2f}.",
        )
        return redirect("payment_page", booking_id=booking.id)

    if verified_paid > Decimal("0.00"):
        payment_type = "balance"
        expected_amount = remaining_balance.quantize(Decimal("0.01"))
    else:
        if payment_option not in {"downpayment", "full"}:
            messages.error(request, "Please select Downpayment or Full Payment.")
            return redirect("payment_page", booking_id=booking.id)
        if payment_option == "downpayment":
            payment_type = "downpayment"
            expected_amount = dp_amount.quantize(Decimal("0.01"))
        else:
            payment_type = "full"
            expected_amount = total_price.quantize(Decimal("0.01"))

    if amount.quantize(Decimal("0.01")) != expected_amount:
        messages.error(
            request,
            f"Invalid amount for selected option. Expected PHP {expected_amount:,.2f}.",
        )
        return redirect("payment_page", booking_id=booking.id)

    if payment_type == "downpayment" and amount >= total_price:
        messages.error(
            request,
            "Downpayment amount cannot be the same as full payment amount.",
        )
        return redirect("payment_page", booking_id=booking.id)

    if payment_type == "full" and amount < total_price and verified_paid <= Decimal("0.00"):
            messages.error(
                request,
                "Full payment must match the total booking amount.",
            )
            return redirect("payment_page", booking_id=booking.id)

    amount_cents = int(amount * 100)
    success_url = request.build_absolute_uri(reverse("paymongo_success", kwargs={"booking_id": booking.id}))
    cancel_url = request.build_absolute_uri(reverse("paymongo_cancel", kwargs={"booking_id": booking.id}))

    paymongo_type = "gcash"

    checkout_session = create_paymongo_checkout_session(
        amount=amount_cents,
        booking_id=booking.id,
        success_url=success_url,
        cancel_url=cancel_url,
        payment_type=paymongo_type,
        description=f"Payment for Booking #{booking.id}"
    )

    if not checkout_session:
        messages.error(request, "Failed to create payment session. Please try again.")
        return redirect("payment_page", booking_id=booking.id)

    try:
        session_id = checkout_session["data"]["id"]
        checkout_url = checkout_session["data"]["attributes"]["checkout_url"]
    except (KeyError, TypeError):
        messages.error(request, "Invalid response from payment provider.")
        return redirect("payment_page", booking_id=booking.id)

    transaction_ref = str(uuid.uuid4()).replace("-", "")[:16].upper()
    Payment.objects.create(
        booking=booking,
        amount=amount,
        payment_method=f"paymongo_{payment_method_type}",
        payment_type=payment_type,
        payment_status="pending",
        transaction_ref=transaction_ref,
        paymongo_checkout_session_id=session_id,
        paymongo_checkout_url=checkout_url,
    )

    log_action(request.user, f"Created PayMongo checkout session for Booking #{booking.id}.")
    return redirect(checkout_url)


@login_required
def paymongo_success(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    pending_paymongo = (
        booking.payments.filter(
            payment_status="pending",
            payment_method__startswith="paymongo_",
            paymongo_checkout_session_id__gt="",
        )
        .order_by("-created_at")
        .first()
    )

    if not pending_paymongo:
        messages.info(
            request,
            "No pending PayMongo payment found. If you already paid, please wait for confirmation.",
        )
        return redirect("payment_page", booking_id=booking_id)

    checkout_data = retrieve_paymongo_checkout_session(
        pending_paymongo.paymongo_checkout_session_id
    )
    if not checkout_data:
        messages.info(
            request,
            "Payment is still processing. Please refresh this page in a moment.",
        )
        return redirect("payment_page", booking_id=booking_id)

    attrs = checkout_data.get("data", {}).get("attributes", {})
    checkout_status = (attrs.get("status") or "").lower()

    payment_intent = attrs.get("payment_intent") or {}
    payment_intent_id = payment_intent.get("id", "")
    payment_intent_status = (
        payment_intent.get("attributes", {}).get("status", "") or ""
    ).lower()
    if payment_intent_id and payment_intent_status not in {
        "succeeded",
        "paid",
        "failed",
        "cancelled",
    }:
        payment_data = retrieve_paymongo_payment(payment_intent_id)
        payment_intent_status = (
            payment_data.get("data", {}).get("attributes", {}).get("status", "")
            if payment_data
            else payment_intent_status
        )
        payment_intent_status = (payment_intent_status or "").lower()

    if checkout_status in {"paid", "succeeded"} or payment_intent_status in {
        "succeeded",
        "paid",
    }:
        pending_paymongo.paymongo_payment_id = payment_intent_id
        pending_paymongo.notes = "Paid via PayMongo. Awaiting admin verification."
        pending_paymongo.save(
            update_fields=["paymongo_payment_id", "notes"]
        )

        AdminNotification.objects.create(
            booking=booking,
            user=request.user,
            message=(
                f"New PayMongo payment submitted for Booking #{booking.id} "
                f"by {request.user.get_full_name() or request.user.username}. "
                f"Amount: PHP {pending_paymongo.amount:,.2f}"
            ),
        )

        Notification.objects.create(
            user=booking.user,
            booking=booking,
            message=(
                f"Your payment of PHP {pending_paymongo.amount:,.2f} for Booking "
                f"#{booking.id} was received and is now pending admin verification."
            ),
        )
        messages.success(
            request,
            "Payment received! It is now pending admin verification.",
        )
    elif checkout_status in {"failed", "expired", "cancelled"}:
        pending_paymongo.payment_status = "rejected"
        pending_paymongo.notes = "PayMongo checkout did not complete."
        pending_paymongo.save(update_fields=["payment_status", "notes"])
        messages.warning(
            request,
            "Payment did not complete. Please try again.",
        )
    else:
        messages.info(
            request,
            "Payment is still pending confirmation. Please check again shortly.",
        )
    return redirect("payment_page", booking_id=booking_id)


@login_required
def paymongo_cancel(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    pending_paymongo = (
        booking.payments.filter(
            payment_status="pending",
            payment_method__startswith="paymongo_",
            paymongo_checkout_session_id__gt="",
        )
        .order_by("-created_at")
        .first()
    )
    if pending_paymongo:
        pending_paymongo.payment_status = "rejected"
        pending_paymongo.notes = "Customer cancelled PayMongo checkout."
        pending_paymongo.save(update_fields=["payment_status", "notes"])
    messages.info(request, "Payment was cancelled. You can try again at any time.")
    return redirect("payment_page", booking_id=booking_id)


@require_POST
def paymongo_webhook(request):
    payload = request.body.decode('utf-8')
    signature_header = request.headers.get('Paymongo-Signature', '')

    if not verify_paymongo_webhook_signature(payload, signature_header):
        return JsonResponse({"error": "Invalid signature"}, status=400)

    try:
        data = json.loads(payload)
        event_attrs = data.get("data", {}).get("attributes", {})
        event_type = (event_attrs.get("type") or "").strip()
        event_data = event_attrs.get("data", {}) or {}
        event_data_attrs = event_data.get("attributes", {}) or {}
        checkout_session_id = event_data.get("id") or event_data_attrs.get("id")

        if not checkout_session_id:
            return JsonResponse({"success": True})

        payment = Payment.objects.filter(
            paymongo_checkout_session_id=checkout_session_id
        ).first()
        if not payment:
            return JsonResponse({"success": True})

        if event_type == "checkout_session.payment.paid" and payment.payment_status == "pending":
            payment_ref = ""
            payments_arr = event_data_attrs.get("payments") or []
            if payments_arr and isinstance(payments_arr, list):
                payment_ref = payments_arr[0].get("id", "")

            if payment_ref:
                payment.paymongo_payment_id = payment_ref
            payment.notes = "Paid via PayMongo webhook. Awaiting admin verification."
            payment.save(update_fields=["paymongo_payment_id", "notes"])
        elif event_type in {"checkout_session.payment.failed", "checkout_session.expired"}:
            if payment.payment_status == "pending":
                payment.payment_status = "rejected"
                payment.notes = f"PayMongo webhook event: {event_type}"
                payment.save(update_fields=["payment_status", "notes"])

        return JsonResponse({"success": True})
    except Exception as e:
        print(f"PayMongo Webhook Error: {e}")
        return JsonResponse({"error": "Failed to process webhook"}, status=500)
